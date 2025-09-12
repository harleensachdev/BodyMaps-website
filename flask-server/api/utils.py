from flask import Blueprint, send_file, make_response, request, jsonify
from services.nifti_processor import NiftiProcessor
from services.session_manager import SessionManager, generate_uuid
from services.auto_segmentor import run_auto_segmentation
from models.application_session import ApplicationSession
from models.combined_labels import CombinedLabels
from models.base import db
from constants import Constants

from io import BytesIO
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

import os
import uuid
import time
import math
import numpy as np
import nibabel as nib
from scipy.ndimage import distance_transform_edt
from collections import defaultdict
from services.npz_processor import NpzProcessor
from PIL import Image
from openpyxl import load_workbook
# Track last session validation time
last_session_check = datetime.now()

# Progress tracking structure: {session_id: (start_time, expected_total_seconds)}
progress_tracker = {}

def id_is_training(index):
    return index < 9000

def combine_label_npz(index: int):
    npz_processor = NpzProcessor()
    npz_processor.combine_labels(index)
    return
def get_panTS_id(index):
    cur_case_id = str(index)
    iter = max(0, 8 - len(str(index)))
    for _ in range(iter):
        cur_case_id = "0" + cur_case_id
    cur_case_id = "PanTS_" + cur_case_id    
    return cur_case_id

def clean_nan(obj):
    """Recursively replace NaN with None for JSON serialization."""
    if isinstance(obj, dict):
        return {k: clean_nan(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_nan(elem) for elem in obj]
    elif isinstance(obj, float) and math.isnan(obj):
        return None
    else:
        return obj

def format_value(value):
    """Format values for display, replacing 999999 or None with 'N/A'."""
    return "N/A" if value in [999999, None] else str(value)

def organname_to_name(filename):
    """Convert a NIfTI file name to a human-readable organ name."""
    name = filename.replace(".nii.gz", "").replace("_", " ")
    return name.title()

def get_mask_data_internal(id, fallback=False):
    """Retrieve or compute organ metadata from NIfTI and mask paths for a session."""
    try:
        subfolder = "ImageTr" if int(id) < 9000 else "ImageTe"
        label_subfolder = "LabelTr" if int(id) < 9000 else "LabelTe"
        main_nifti_path = f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(id)}/{Constants.MAIN_NIFTI_FILENAME}"
        combined_labels_path = f"{Constants.PANTS_PATH}/data/{label_subfolder}/{get_panTS_id(id)}/{Constants.COMBINED_LABELS_NIFTI_FILENAME}"
        print(f"[INFO] Processing NIFTI for id {id}")
        organ_intensities = None
        
        organ_intensities_path = f"{Constants.PANTS_PATH}/data/{label_subfolder}/{get_panTS_id(id)}/{Constants.ORGAN_INTENSITIES_FILENAME}"
        if not os.path.exists(organ_intensities_path) or not os.path.exists(combined_labels_path):
            npz_processor = NpzProcessor()
            labels, organ_intensities = npz_processor.combine_labels(int(id), keywords={"pancrea": "pancreas"}, save=True)
        else: 
            with open(organ_intensities_path, "r") as f:
                organ_intensities = json.load(f)
        
        nifti_processor = NiftiProcessor(main_nifti_path, combined_labels_path)
        nifti_processor.set_organ_intensities(organ_intensities)
        organ_metadata = nifti_processor.calculate_metrics()
        organ_metadata = clean_nan(organ_metadata)

        return organ_metadata

    except Exception as e:
        print(f"[ERROR] get_mask_data_internal: {e}")
        return {"error": str(e)}

def generate_distinct_colors(n):
    """Generate n visually distinct RGB colors."""
    import colorsys
    HSV_tuples = [(x / n, 0.7, 0.9) for x in range(n)]
    RGB_tuples = [tuple(int(c * 255) for c in colorsys.hsv_to_rgb(*hsv)) for hsv in HSV_tuples]
    return RGB_tuples

def fill_voids_with_nearest_label(label_array):
    """Fill all 0-valued voxels with the nearest non-zero label."""
    mask = label_array == 0
    if not np.any(mask):
        return label_array

    nonzero_coords = np.array(np.nonzero(label_array)).T
    distances, indices = distance_transform_edt(mask, return_indices=True)
    filled_array = label_array.copy()
    filled_array[mask] = label_array[tuple(indices[:, mask])]
    return filled_array

def build_adjacency_graph(label_array):
    """Build adjacency graph of label connectivity in 6 directions."""
    adjacency = defaultdict(set)
    offsets = [(-1, 0, 0), (1, 0, 0),
               (0, -1, 0), (0, 1, 0),
               (0, 0, -1), (0, 0, 1)]

    for dx, dy, dz in offsets:
        shifted = np.roll(label_array, shift=(dx, dy, dz), axis=(0, 1, 2))
        mask = (label_array != shifted) & (label_array != 0) & (shifted != 0)
        l1 = label_array[mask]
        l2 = shifted[mask]
        for a, b in zip(l1, l2):
            if a != b:
                adjacency[a].add(b)
                adjacency[b].add(a)
    return adjacency

def assign_colors_with_high_contrast(label_ids, adjacency_graph, min_initial_colors=20, max_total_colors=50):
    """
    Assign colors to labels such that adjacent labels have different colors,
    maximizing contrast and balance.
    """
    from itertools import combinations
    import colorsys

    def generate_distinct_colors(n):
        HSV_tuples = [(x / n, 0.7, 0.9) for x in range(n)]
        RGB_tuples = [tuple(int(c * 255) for c in colorsys.hsv_to_rgb(*hsv)) for hsv in HSV_tuples]
        return RGB_tuples

    def can_use_color(label, color_idx, assignments, adjacency_graph):
        for neighbor in adjacency_graph[label]:
            if assignments.get(neighbor) == color_idx:
                return False
        return True

    label_ids = sorted(label_ids)
    assignments = {}
    num_colors = min_initial_colors
    color_usage_count = {i: 0 for i in range(num_colors)}

    while True:
        colors = generate_distinct_colors(num_colors)
        assignments.clear()
        color_usage_count = {i: 0 for i in range(num_colors)}
        success = True

        for label in label_ids:
            color_order = sorted(range(num_colors), key=lambda c: (color_usage_count[c], c))
            for color_idx in color_order:
                if can_use_color(label, color_idx, assignments, adjacency_graph):
                    assignments[label] = color_idx
                    color_usage_count[color_idx] += 1
                    break
            else:
                success = False
                break

        if success:
            break
        elif num_colors >= max_total_colors:
            print(f"⚠️ Warning: reached max color count {max_total_colors}, some neighbors may share color")
            break
        else:
            num_colors += 1

    final_colors = generate_distinct_colors(num_colors)
    print(f"✅ Final color count used: {len(set(assignments.values()))}")

    color_map = {
        str(round(label)): {
            "R": final_colors[color_idx][0],
            "G": final_colors[color_idx][1],
            "B": final_colors[color_idx][2],
            "A": 128
        }
        for label, color_idx in assignments.items()
    }

    return color_map, color_usage_count

def wait_for_file(filepath, timeout=30, check_interval=0.5):
    """Wait until a file exists, or timeout is reached."""
    start_time = time.time()
    while not os.path.exists(filepath):
        if time.time() - start_time > timeout:
            raise TimeoutError(f"Timeout: File {filepath} not found after {timeout} seconds.")
        time.sleep(check_interval)

def volume_to_png(volume, axis=2, index=None):
    if index is None:
        index = volume.shape[axis] // 2
    
    slice_ = np.take(volume, index, axis=axis)
    # window_center = 40 
    # window_width = 400 
    # min_val = window_center - window_width / 2
    # max_val = window_center + window_width / 2

    # slice_clipped = np.clip(slice_, min_val, max_val)
    # slice_norm = 255 * (slice_clipped - min_val) / (max_val - min_val)
    slice_norm = 255 * (slice_ - np.min(slice_)) / (np.max(slice_) - np.min(slice_))
    slice_norm = slice_norm.astype(np.uint8)
    
    slice_norm = np.rot90(slice_norm, k=1)
    slice_norm = np.flip(slice_norm, axis=0)

    pil_img = Image.fromarray(slice_norm)
    buf = BytesIO()
    pil_img.save(buf, format="PNG")
    buf.seek(0)
    return buf
def generate_pdf_with_template(
    output_pdf,
    folder_name,
    ct_path,
    mask_path,
    template_pdf,
    temp_pdf_path,
    id,
    extracted_data=None,
    column_headers=None,
):
    import os
    import nibabel as nib
    import numpy as np
    import pandas as pd
    from PyPDF2 import PdfReader, PdfWriter
    from PyPDF2._page import PageObject
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter

    LABELS = {v: k for k, v in Constants.PREDEFINED_LABELS.items()}
    NAME_TO_ORGAN = {
        # Pancreas and its lesions
        "pancreas": "pancreas",
        "pancreas_body": "pancreas",
        "pancreas_head": "pancreas",
        "pancreas_tail": "pancreas",
        "pancreatic_lesion": "pancreas",
        "pancreatic_duct": "pancreas",

        # All other organs: map to self
        "aorta": "aorta",
        "adrenal_gland_left": "adrenal_gland_left",
        "adrenal_gland_right": "adrenal_gland_right",
        "bladder": "bladder",
        "common_bile_duct": "common_bile_duct",
        "celic_artery": "celiac_artery",
        "colon": "colon",
        "duodenum": "duodenum",
        "femur_right": "femur_right",
        "femur_left": "femur_left",
        "gall_bladder": "gall_bladder",
        "postcava": "postcava",
        "kidney_left": "kidney_left",
        "kidney_right": "kidney_right",
        "liver": "liver",
        "postcava": "postcava",
        "prostate": "prostate",
        "superior_mesenteric_artery": "superior_mesenteric_artery",
        "intestine": "intestine",
        "spleen": "spleen",
        "stomach": "stomach",
        "veins": "veins",
    }

    try:
        temp_pdf = canvas.Canvas(temp_pdf_path, pagesize=letter)
        width, height = letter
        left_margin, top_margin = 50, 100
        line_height, section_spacing = 12, 30
        y_position = height - top_margin

        def reset_page():
            nonlocal y_position
            temp_pdf.showPage()
            y_position = height - 120
            temp_pdf.setFont("Helvetica", 10)

        def write_wrapped_text(x, y, content, bold=False, font_size=10, max_width=None):
            temp_pdf.setFont("Helvetica-Bold" if bold else "Helvetica", font_size)
            words = content.split()
            current_line = ""
            max_width = max_width or width - left_margin * 2
            for word in words:
                if temp_pdf.stringWidth(current_line + word + " ", "Helvetica", font_size) > max_width:
                    temp_pdf.drawString(x, y, current_line.strip())
                    y -= line_height
                    current_line = f"{word} "
                    if y < 50:
                        reset_page()
                        y = y_position
                else:
                    current_line += f"{word} "
            if current_line:
                temp_pdf.drawString(x, y, current_line.strip())
                y -= line_height
            return y

        def safe_extract(index, default="N/A"):
            if extracted_data is not None and index in extracted_data:
                val = extracted_data[index]
                return "N/A" if pd.isna(val) else val
            return default
        
        wb = load_workbook(os.path.join(Constants.PANTS_PATH, "data", "metadata.xlsx"))
        sheet = wb["PanTS_metadata"]
        age = None
        sex = "-"
        contrast = ""
        study_detail = ""
        for row in sheet.iter_rows(values_only=True):
            if row[0] == get_panTS_id(folder_name):
                age = row[5]
                sex = row[4]
                contrast = row[3]
                study_detail = row[8]
                break

        # Title
        temp_pdf.setFont("Helvetica-Bold", 26)
        title_text = "MEDICAL REPORT"
        title_width = temp_pdf.stringWidth(title_text, "Helvetica-Bold", 26)
        temp_pdf.drawString((width - title_width) / 2, height - 70, title_text)
        y_position = height - 100

        # Patient info
        temp_pdf.setFont("Helvetica-Bold", 12)
        temp_pdf.drawString(left_margin, y_position, "PATIENT INFORMATION")
        y_position -= line_height

        left_y = write_wrapped_text(left_margin, y_position, f"PANTS ID: {folder_name}")
        right_y = write_wrapped_text(width / 2, y_position, f"Sex: {sex}")
        y_position -= line_height
        
        write_wrapped_text(left_margin, y_position, f"Age: {age}")
        
        y_position = min(left_y, right_y) - section_spacing

        # Imaging detail
        temp_pdf.setFont("Helvetica-Bold", 12)
        temp_pdf.drawString(left_margin, y_position, "IMAGING DETAIL")
        y_position -= line_height

        ct_nii = nib.load(ct_path)
        spacing = ct_nii.header.get_zooms()
        shape = ct_nii.shape

        try:
            scanner_info = str(ct_nii.header['descrip'].tobytes().decode('utf-8')).strip().replace('\x00', '')
        except Exception:
            scanner_info = "N/A"


        y_position = write_wrapped_text(left_margin, y_position, f"Spacing: {spacing}")
        y_position = write_wrapped_text(left_margin, y_position, f"Shape: {shape}")
        y_position = write_wrapped_text(left_margin, y_position, f"Study type: {study_detail}")
        y_position = write_wrapped_text(left_margin, y_position, f"Contrast: {contrast}")
        y_position -= section_spacing

        # Load image data
        ct_array = ct_nii.get_fdata()
        mask_array = nib.load(mask_path).get_fdata().astype(np.uint8)
        voxel_volume = np.prod(nib.load(mask_path).header.get_zooms()) / 1000  # mm³ to cm³
        print(np.unique(mask_array))

        # AI Measurements
        temp_pdf.setFont("Helvetica-Bold", 12)
        temp_pdf.drawString(left_margin, y_position, "AI MEASUREMENTS")
        y_position -= line_height

        # Table configuration
        headers = ["Organ", "Volume (cc)", "Mean HU"]
        col_widths = [120, 100, 100]
        row_height = 20

        def draw_table_row(row_data, is_header=False):
            nonlocal y_position
            if y_position - row_height < 50:
                reset_page()
                temp_pdf.setFont("Helvetica-Bold", 12)
                temp_pdf.drawString(left_margin, y_position, "AI MEASUREMENTS (continued)")
                y_position -= line_height
                draw_table_row(headers, is_header=True)
            x = left_margin
            temp_pdf.setFont("Helvetica-Bold" if is_header else "Helvetica", 9)
            for i, cell in enumerate(row_data):
                temp_pdf.drawString(x + 2, y_position - row_height + 5, str(cell))
                temp_pdf.line(x, y_position, x, y_position - row_height)
                x += col_widths[i]
            temp_pdf.line(left_margin + sum(col_widths), y_position, left_margin + sum(col_widths), y_position - row_height)
            temp_pdf.line(left_margin, y_position, left_margin + sum(col_widths), y_position)
            y_position -= row_height
            temp_pdf.line(left_margin, y_position, left_margin + sum(col_widths), y_position)

        draw_table_row(headers, is_header=True)

        lession_volume_dict={}
        for organ, label_id in LABELS.items():
            if organ in NAME_TO_ORGAN and NAME_TO_ORGAN[organ] != organ:
                mask = (mask_array == label_id)
                if not np.any(mask):
                    print("none")
                    continue
                volume = np.sum(mask) * voxel_volume
                mean_hu = np.mean(ct_array[mask])
                if NAME_TO_ORGAN[organ] in lession_volume_dict:
                    lession_volume_dict[NAME_TO_ORGAN[organ]]["number"] += 1
                    lession_volume_dict[NAME_TO_ORGAN[organ]]["volume"] += volume
                else:
                    lession_volume_dict[NAME_TO_ORGAN[organ]] = {
                        "number": 1,
                        "volume": volume
                    }
                    
        print(lession_volume_dict)
        
        for organ, label_id in LABELS.items():
            if organ in NAME_TO_ORGAN and NAME_TO_ORGAN[organ] != organ:
                continue
            if label_id == 0:
                continue
            mask = (mask_array == label_id)
            if not np.any(mask):
                continue
            volume = np.sum(mask) * voxel_volume
            mean_hu = np.mean(ct_array[mask])
            
            if organ in lession_volume_dict:
                row = [organ.replace('_', ' '), f"{volume:.2f}", f"{mean_hu:.1f}"]
            else:
                row = [organ.replace('_', ' '), f"{volume:.2f}", f"{mean_hu:.1f}"]
            draw_table_row(row)

        # y_position -= section_spacing

        # === Step 2: PDAC Staging ===
        # temp_pdf.setFont("Helvetica-Bold", 12)
        # temp_pdf.drawString(left_margin, y_position, "PDAC STAGING")
        # y_position -= line_height

        # try:
        #     pdac_info = get_pdac_staging(id)
        #     print(pdac_info, id)
        #     pdac_text = pdac_info.get("staging_description", "No staging data available.")
        # except Exception:
        #     pdac_text = "Error fetching PDAC staging information."

        # y_position = write_wrapped_text(left_margin, y_position, pdac_text, bold=False, font_size=10)
        # === Step 3: Key Images ===
        
        # include_liver = np.count_nonzero(mask_array == LABELS["liver"]) > 0
        # include_pancreas = lession_volume_dict.get("pancreas", {}).get("number", 0) > 0
        # include_kidney = np.count_nonzero(mask_array == LABELS["kidney_left"]) > 0 or np.count_nonzero(mask_array == LABELS["kidney_right"]) > 0
        # print(include_liver, include_pancreas, include_kidney)
        # if include_liver or include_pancreas or include_kidney:
        #     def check_and_reset_page(space_needed):
        #         nonlocal y_position
        #         if y_position - space_needed < 50:
        #             reset_page()

        #     temp_pdf.showPage()
        #     y_position = height - top_margin
        #     temp_pdf.setFont("Helvetica-Bold", 14)
        #     # temp_pdf.drawString(left_margin, y_position, "KEY IMAGES")
        #     y_position -= section_spacing

        #     organs = {
        #         "liver": include_liver,
        #         "pancreas": include_pancreas,
        #         "kidney_left": include_kidney,
        #         "kidney_right": include_kidney
        #     }
            # download_clean_folder(ct_path.replace("/inputs/", "/outputs/").rsplit("/", 1)[0])
            # for organ in organs:
            #     organ_data = lession_volume_dict.get(organ)
            #     if not organ_data or organ_data.get("number", 0) == 0:
            #         continue

            #     header = f"{organ.replace('_', ' ').upper()} TUMORS"
            #     check_and_reset_page(line_height)
            #     temp_pdf.setFont("Helvetica", 12)
            #     temp_pdf.drawString(left_margin, y_position, header)
            #     y_position -= line_height
            #     print(organ, organ_data)
            #     check_and_reset_page(220)
            #     overlay_path = f"/tmp/{organ}_overlay.png"
            #     print(ct_path, mask_path)
            #     organ_mask_path = mask_path.replace('combined_labels.nii.gz', 'segmentations/'+organ+'.nii.gz')
            #     print(organ_mask_path)
            #     if create_overlay_image(ct_path, organ_mask_path, overlay_path, color="red"):
            #         try:
            #             temp_pdf.drawImage(overlay_path, left_margin, y_position - 200, width=200, height=200)
            #         except:
            #             print(overlay_path)
            #     check_and_reset_page(220)
            #     zoom_path = f"/tmp/{organ}_zoomed.png"
            #     if zoom_into_labeled_area(ct_path, organ_mask_path, zoom_path, color="red"):
            #         temp_pdf.drawImage(zoom_path, left_margin + 250, y_position - 205, width=210, height=210)
            #     print('521')
            #     y_position -= 220

        temp_pdf.save()

        # Merge with template
        template_reader =  PdfReader(template_pdf)
        content_reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        for page in content_reader.pages:
            template_page = template_reader.pages[0]
            merged_page = PageObject.create_blank_page(
                width=template_page.mediabox.width,
                height=template_page.mediabox.height
            )
            merged_page.merge_page(template_page)
            merged_page.merge_page(page)
            writer.add_page(merged_page)

        with open(output_pdf, "wb") as f:
            writer.write(f)

                
    except Exception as e:
        raise RuntimeError(f"Error generating PDF for {folder_name}: {e}")
    finally:
        if os.path.exists(temp_pdf_path):
            os.remove(temp_pdf_path)

# Helper Function to Process CT and Mask
def get_most_labeled_slice(ct_path, mask_path, output_png, contrast_min=-150, contrast_max=250):
    """
    Load CT and mask, ensure RAS orientation, find the most labeled slice, and generate an overlay image.
    """

    try:
        import SimpleITK as sitk
        import matplotlib
        matplotlib.use('Agg')  # ✅ 关键：不再尝试调用 GUI

        import matplotlib.pyplot as plt

        # Load the CT scan and mask
        ct_scan = sitk.ReadImage(ct_path)
        print('543',mask_path)
        mask = sitk.ReadImage(mask_path)
        print(mask_path)
        # Reorient to RAS
        ct_scan = sitk.DICOMOrient(ct_scan, 'RAS')
        mask = sitk.DICOMOrient(mask, 'RAS')

        # Convert to numpy arrays
        ct_array = sitk.GetArrayFromImage(ct_scan)
        mask_array = sitk.GetArrayFromImage(mask)

        # Check for shape mismatches
        if ct_array.shape != mask_array.shape:
            raise ValueError(f"Shape mismatch: CT shape {ct_array.shape}, Mask shape {mask_array.shape}")

        # Find the slice with the most labels
        slice_sums = np.sum(mask_array, axis=(1, 2))
        most_labeled_slice_index = np.argmax(slice_sums)

        # Get the CT and mask slices
        ct_slice = ct_array[most_labeled_slice_index]
        mask_slice = mask_array[most_labeled_slice_index]

        # Apply mirroring
        ct_slice = np.fliplr(ct_slice)
        mask_slice = np.fliplr(mask_slice)

        # Apply contrast adjustment
        ct_slice = np.clip(ct_slice, contrast_min, contrast_max)
        ct_slice = (ct_slice - contrast_min) / (contrast_max - contrast_min) * 255
        ct_slice = ct_slice.astype(np.uint8)

        # Overlay mask contours on CT slice
        plt.figure(figsize=(6, 6))
        plt.imshow(ct_slice, cmap='gray', origin='lower')
        plt.contour(mask_slice, colors='red', linewidths=1)  # Use red contours for the mask
        plt.axis('off')
        plt.savefig(output_png, bbox_inches="tight", pad_inches=0)
        plt.close()
        print('586')
        return True
    except:
        return False

def create_overlay_image(ct_path, mask_path, output_path, color="red"):
    """
    Generate overlay images for most labeled slices using the unified RAS orientation logic.
    """
    return get_most_labeled_slice(ct_path, mask_path, output_path)


# Helper Function to Zoom into Labeled Area
def zoom_into_labeled_area(ct_path, mask_path, output_path, color="red"):
    """
    Create a zoomed-in view of the largest labeled area with consistent RAS orientation.
    """
    import SimpleITK as sitk
    import matplotlib.pyplot as plt
    try:
        # Load the CT scan and mask
        ct_scan = sitk.ReadImage(ct_path)
        mask = sitk.ReadImage(mask_path)

        # Reorient to RAS
        ct_scan = sitk.DICOMOrient(ct_scan, 'RAS')
        mask = sitk.DICOMOrient(mask, 'RAS')

        # Convert to numpy arrays
        ct_array = sitk.GetArrayFromImage(ct_scan)
        mask_array = sitk.GetArrayFromImage(mask)

        # Check for shape mismatches
        if ct_array.shape != mask_array.shape:
            raise ValueError(f"Shape mismatch: CT shape {ct_array.shape}, Mask shape {mask_array.shape}")

        # Find the slice with the most labels
        slice_sums = np.sum(mask_array, axis=(1, 2))
        largest_slice_idx = np.argmax(slice_sums)
        if slice_sums[largest_slice_idx] == 0:
            raise ValueError("No labeled area found in the mask.")

        # Get the mask slice and calculate the bounding box
        mask_slice = mask_array[largest_slice_idx]
        coords = np.array(np.where(mask_slice))
        min_row, max_row = np.min(coords[0]), np.max(coords[0])
        min_col, max_col = np.min(coords[1]), np.max(coords[1])
        padding = 20
        min_row = max(min_row - padding, 0)
        max_row = min(max_row + padding, mask_slice.shape[0])
        min_col = max(min_col - padding, 0)
        max_col = min(max_col + padding, mask_slice.shape[1])

        # Extract the zoomed region
        zoomed_image = ct_array[largest_slice_idx][min_row:max_row, min_col:max_col]
        zoomed_mask = mask_array[largest_slice_idx][min_row:max_row, min_col:max_col]

        # Apply mirroring
        zoomed_image = np.fliplr(zoomed_image)
        zoomed_mask = np.fliplr(zoomed_mask)

        # Apply contrast adjustment to the zoomed CT slice
        zoomed_image = np.clip(zoomed_image, -150, 250)
        zoomed_image = (zoomed_image + 150) / 400 * 255
        zoomed_image = zoomed_image.astype(np.uint8)

        # Save the zoomed-in image with overlay
        plt.figure(figsize=(6, 6))
        plt.imshow(zoomed_image, cmap="gray", origin="lower")
        plt.contour(zoomed_mask, colors=color, linewidths=1)
        plt.axis("off")
        plt.savefig(output_path, bbox_inches="tight")
        plt.close()
        return True
    except Exception as e:
        return False

def get_pdac_staging(clabel_id):
    try:
        subfolder = "ImageTr" if int(clabel_id) < 9000 else "ImageTe"
        label_subfolder = "LabelTr" if int(clabel_id) < 9000 else "LabelTe"
        main_nifti_path = f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(clabel_id)}/{Constants.MAIN_NIFTI_FILENAME}"
        combined_labels_path = f"{Constants.PANTS_PATH}/data/{label_subfolder}/{get_panTS_id(clabel_id)}/{Constants.COMBINED_LABELS_NIFTI_FILENAME}"
        
        nifti_processor = NiftiProcessor(main_nifti_path, combined_labels_path)
        staging_result = nifti_processor.calculate_pdac_sma_staging()

        return {"staging_description": staging_result}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": f"PDAC staging failed: {str(e)}"}

import json
def download_clean_folder(root):
    """
    如果文件正好匹配4个目标名，则删除其中两个，并将combined_labels.nii.gz根据dataset.json分割为独立器官文件。
    """
    target_files = {
        "combined_labels.nii.gz",
        "dataset.json",
        "plans.json",
        "predict_from_raw_data_args.json"
    }

    actual_files = set(os.listdir(root))
    if actual_files == target_files:
        # 删除 plans.json 和 predict_from_raw_data_args.json
        for fname in ["plans.json", "predict_from_raw_data_args.json"]:
            fpath = os.path.join(root, fname)
            if os.path.exists(fpath):
                os.remove(fpath)
                print(f"🗑️ Removed during zip: {fpath}")

        # 读取 dataset.json
        dataset_json_path = os.path.join(root, "dataset.json")
        with open(dataset_json_path, 'r') as f:
            dataset_info = json.load(f)

        labels = dataset_info["labels"]  # 获取标签名与ID的映射

        # 读取 combined_labels.nii.gz
        combined_path = os.path.join(root, "combined_labels.nii.gz")
        combined_img = nib.load(combined_path)
        combined_data = combined_img.get_fdata()
        affine = combined_img.affine

        # 创建 segmentations 文件夹
        seg_folder = os.path.join(root, "segmentations")
        os.makedirs(seg_folder, exist_ok=True)

        # 为每个标签生成单独的 mask 文件
        for label_name, label_value in labels.items():
            mask = (combined_data == label_value).astype(np.uint8)
            label_img = nib.Nifti1Image(mask, affine)
            out_path = os.path.join(seg_folder, f"{label_name}.nii.gz")
            nib.save(label_img, out_path)
            print(f"✅ Saved: {out_path}")
        os.remove(dataset_json_path)
    else:
        print("ℹ️ Folder content does not match the expected file set. Skipping cleanup and split.")