from flask import Blueprint, send_file, make_response, request, jsonify, Response
from werkzeug.utils import secure_filename
from services.nifti_processor import NiftiProcessor
from services.session_manager import SessionManager, generate_uuid
from services.auto_segmentor import run_auto_segmentation, cancel_all_inference
from services.mesh_generation import generate_mesh_manifest, generate_organ_glb_bytes
from services.inference_job_queue import InferenceJobQueue
from services.intent_parser import parse_intent
from services.ollama_client import (
    DEFAULT_OLLAMA_MODEL,
    OllamaUnavailable,
    chat_json,
    list_ollama_models,
)
from services.segmentation_metrics import calculate_session_metrics
from models.application_session import ApplicationSession
from models.combined_labels import CombinedLabels
from models.base import db
from constants import Constants
import zipfile
import json
import pandas as pd

from pathlib import Path
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm

from sqlalchemy.orm import aliased
import os
import io
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
import nibabel as nib
import uuid

from datetime import datetime, timedelta
from .utils import *
import requests  # ⭐ 只在這裡 import 一次 requests

# 建立 blueprint
api_blueprint = Blueprint("api", __name__)
last_session_check = datetime.now()

import hmac
import threading

# Session/case ids come straight from client requests and are joined into
# filesystem paths below. _is_safe_id (pure, unit-tested in tests/unit/
# test_path_safety.py) rejects anything that could escape the intended
# directory before it touches os.path; secure_filename is the barrier at each
# path-construction site.
from .path_safety import is_safe_id as _is_safe_id


def _load_metadata_cache():
    try:
        xlsx_path = os.path.join(Constants.PANTS_PATH, "metadata.xlsx")
        df = pd.read_excel(xlsx_path, engine="openpyxl")
        cache = {}
        for _, row in df.iterrows():
            pid = str(row.iloc[0])
            cache[pid] = {
                "sex": row.iloc[4] if pd.notna(row.iloc[4]) else "",
                "age": row.iloc[5] if pd.notna(row.iloc[5]) else "",
                "tumor": int(row.iloc[13]) if pd.notna(row.iloc[13]) else 0,
            }
        return cache
    except Exception:
        return {}

_METADATA_CACHE = _load_metadata_cache()

progress_tracker = {}  # {session_id: (start_time, expected_total_seconds)}

INFERENCE_QUEUE_DIR = os.getenv(
    "INFERENCE_QUEUE_DIR",
    os.path.join(Constants.SESSIONS_DIR_NAME, "inference_queue"),
)
inference_job_queue = InferenceJobQueue(INFERENCE_QUEUE_DIR)


def _worker_api_token():
    return (os.getenv("WORKER_API_TOKEN", "") or "").strip()


def _get_worker_id():
    return (
        request.headers.get("X-Worker-Id")
        or request.args.get("worker_id")
        or (request.get_json(silent=True) or {}).get("worker_id")
        or "worker-unknown"
    )


def _require_worker_auth():
    expected = _worker_api_token()
    if not expected:
        return jsonify({"error": "WORKER_API_TOKEN is not configured on server"}), 500

    provided = (request.headers.get("X-Worker-Token", "") or "").strip()
    if not hmac.compare_digest(provided, expected):
        return jsonify({"error": "Unauthorized worker token"}), 401
    return None


def _api_prefix_path() -> str:
    base_path = (Constants.BASE_PATH or "/").rstrip("/")
    return f"{base_path}/api" if base_path else "/api"


def _absolute_api_url(path_suffix: str) -> str:
    root = request.url_root.rstrip("/")
    return f"{root}{_api_prefix_path()}{path_suffix}"


def _public_job_payload(job: dict) -> dict:
    response = {
        "job_id": job.get("job_id"),
        "session_id": job.get("session_id"),
        "model": job.get("model"),
        "status": job.get("status"),
        "attempts": job.get("attempts"),
        "max_attempts": job.get("max_attempts"),
        "error": job.get("error"),
        "created_at": job.get("created_at"),
        "updated_at": job.get("updated_at"),
    }
    if job.get("status") == "succeeded":
        response["download_url"] = _absolute_api_url(f"/jobs/{job.get('job_id')}/download")
    return response


@api_blueprint.route("/proxy-image")
def proxy_image():
    """
    Proxy image requests so the browser only talks to our own origin.
    Front-end will call: /api/proxy-image?url=<encoded_hf_url>
    """
    raw_url = request.args.get("url")
    if not raw_url:
        return Response("Missing url parameter", status=400)

    # 可選安全限制：只允許 HuggingFace 來源
    if not raw_url.startswith("https://huggingface.co/"):
        return Response("Forbidden", status=403)

    try:
        r = requests.get(raw_url, timeout=10)
    except Exception as e:
        return Response(f"Upstream error: {e}", status=502)

    if not r.ok:
        return Response(f"Upstream status {r.status_code}", status=r.status_code)

    content_type = r.headers.get("Content-Type", "image/jpeg")

    resp = Response(r.content, status=200, mimetype=content_type)

    # ⭐ 避免 COEP 再擋圖片
    resp.headers["Cross-Origin-Resource-Policy"] = "cross-origin"
    # resp.headers['Cross-Origin-Opener-Policy'] = 'same-origin'
    # resp.headers['Cross-Origin-Embedder-Policy'] = 'require-corp'

    return resp



from flask import request, jsonify
import numpy as np
import nibabel as nib
from scipy.ndimage import distance_transform_edt, label
from collections import defaultdict
from constants import Constants
import os
from openpyxl import load_workbook


SESSIONS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "tmp")
PDF_DIR = f"{Constants.PERMISSIONS_DIR}/pdf"
os.makedirs(SESSIONS_DIR, exist_ok=True)
os.makedirs(PDF_DIR, exist_ok=True)

def _arg(name: str, default=None):
    return request.args.get(name, default)

@api_blueprint.route('/get_preview/<clabel_ids>', methods=['GET'])
def get_preview(clabel_ids):
    clabel_ids = clabel_ids.split(",")
    res = {}
    for clabel_id in clabel_ids:
        pid = get_panTS_id(clabel_id)
        entry = _METADATA_CACHE.get(pid, {"sex": "", "age": "", "tumor": 0})
        res[clabel_id] = entry
    return jsonify(res)

# if not preloaded
@api_blueprint.route('/get_image_preview/<clabel_id>', methods=['GET'])
def get_image_preview(clabel_id):
    if not _is_safe_id(clabel_id):
        return jsonify({"error": "Invalid id"}), 400
    path = os.path.join(Constants.PANTS_PATH, "profile_only", get_panTS_id(secure_filename(clabel_id)), "profile.jpg")
    if not os.path.exists(path):
        return jsonify({"error": f"File not found: {path} "}), 404
    return send_file(
        path,
        mimetype="image/jpg",   
        as_attachment=False,
        download_name=f"{clabel_id}_slice.jpg"
    )


@api_blueprint.route("/cases/<case_id>/mesh-manifest")
def get_mesh_manifest(case_id):
    if not _is_safe_id(case_id):
        return jsonify({"error": "Invalid id"}), 400
    manifest_path = os.path.join(Constants.MESH_PATH, get_panTS_id(secure_filename(case_id)), "manifest.json")

    if not os.path.exists(manifest_path):
        return jsonify({"error": f"File not found: {manifest_path} "}), 404

    with open(manifest_path, "r", encoding="utf-8") as f:
        manifest = json.load(f)

    return jsonify(manifest)

@api_blueprint.route("/cases/<display_id>/render_only/<filename>")
def get_mesh_file(display_id, filename):
    mesh_path = os.path.join(Constants.MESH_PATH, display_id, filename)
    try:
        response = send_file(
            mesh_path,
            mimetype="model/gltf-binary",
            conditional=False,
        )

    except Exception as e:
        return jsonify({"error": f"Error generating GLB: {str(e)}"}), 500
        

    return response

@api_blueprint.route('/get-label-colormap/<clabel_id>', methods=['GET'])
def get_label_colormap(clabel_id):
    
    clabel_path = os.path.join(Constants.PANTS_PATH, "mask_only", get_panTS_id(int(clabel_id)),  'combined_labels.nii.gz')

    try:
        clabel_array = nib.load(clabel_path)
        clabel_array = clabel_array.get_fdata()
        print("[DEBUG] Nifti loaded, shape =", clabel_array.shape)

        filled_array = fill_voids_with_nearest_label(clabel_array)
        print("[DEBUG] fill_voids_with_nearest_label done")

        adjacency = build_adjacency_graph(filled_array)
        print("[DEBUG] build_adjacency_graph done")

        unique_labels = sorted(adjacency.keys())
        color_map, color_usage_count = assign_colors_with_high_contrast(unique_labels, adjacency)
        print("[DEBUG] Color map generated:", color_map, color_usage_count)

        return jsonify(color_map)

    except Exception as e:
        print("[❌ EXCEPTION]", str(e))
        return jsonify({"error": str(e)}), 500




# @api_blueprint.before_request
# def before_request():
#     global last_session_check
#     current_time = datetime.now()
#     if current_time >= last_session_check + timedelta(minutes=Constants.SCHEDULED_CHECK_INTERVAL):
#         session_manager = SessionManager.instance()
#         expired = session_manager.get_expired()
#         for app_session in expired:
#             session_manager.terminate_session(app_session.session_id)
        
#         last_session_check = current_time

@api_blueprint.route('/', methods=['GET'])
def home():
    return "api"


@api_blueprint.route('/upload', methods=['POST'])
def upload():
    try:
        session_id = request.form.get('SESSION_ID')
        if not session_id:
            return jsonify({"error": "No session ID provided"}), 400
        if not _is_safe_id(session_id):
            return jsonify({"error": "Invalid session ID"}), 400

        base_path = os.path.join(Constants.SESSIONS_DIR_NAME, session_id)
        os.makedirs(base_path, exist_ok=True)

        nifti_multi_dict = request.files
        filenames = list(nifti_multi_dict)
        main_nifti = nifti_multi_dict.get(Constants.MAIN_NIFTI_FORM_NAME)

        if main_nifti:
            main_nifti_path = os.path.join(base_path, Constants.MAIN_NIFTI_FILENAME)
            main_nifti.save(main_nifti_path)
            filenames.remove(Constants.MAIN_NIFTI_FORM_NAME)
        else:
            return jsonify({"error": "Main NIFTI file missing"}), 400

        nifti_processor = NiftiProcessor.from_clabel_path(os.path.join(base_path, Constants.COMBINED_LABELS_FILENAME))

        combined_labels, organ_intensities = nifti_processor.combine_labels(filenames, nifti_multi_dict, save=True)

        resp = {
            'status': "200",
            'session_id': session_id,
            'organ_intensities': organ_intensities
        }
        return jsonify(resp)
    except Exception as e:
        print(f"❌ [Upload Error] {e}")
        return jsonify({"error": "Internal server error"}), 500

@api_blueprint.route('/mask-data', methods=['POST'])
def get_mask_data():
    session_key = request.form.get('sessionKey')
    if not session_key:
        return jsonify({"error": "Missing sessionKey"}), 400

    result = get_mask_data_internal(session_key)
    return jsonify(result)

  
@api_blueprint.route('/get-main-nifti/<clabel_id>.nii.gz', methods=['GET'])
def get_main_nifti(clabel_id):
    if not _is_safe_id(clabel_id):
        return jsonify({"error": "Invalid id"}), 400
    case_dir = f"{Constants.PANTS_PATH}/image_only/{get_panTS_id(secure_filename(clabel_id))}"
    main_nifti_path = f"{case_dir}/{Constants.MAIN_NIFTI_FILENAME}"

    # ?res=low → serve the precomputed low-res copy when present (much smaller/faster
    # for big full-body scans). Falls back to full res if it hasn't been generated.
    if (request.args.get('res') or '').strip().lower() == 'low':
        low_path = f"{case_dir}/{Constants.MAIN_NIFTI_FILENAME.replace('.nii.gz', '_lowres.nii.gz')}"
        if os.path.exists(low_path):
            main_nifti_path = low_path

    if os.path.exists(main_nifti_path):
        response = make_response(send_file(main_nifti_path, mimetype='application/gzip'))

        response.headers['Cross-Origin-Resource-Policy'] = 'cross-origin'
        # response.headers['Cross-Origin-Embedder-Policy'] = 'require-corp'
        # Volumes are immutable per case — let the browser cache so revisits are instant.
        response.headers['Cache-Control'] = 'public, max-age=604800, immutable'

    else:
        print(f"Could not find filepath: {main_nifti_path}. ")
        return jsonify({"error": "Could not find filepath"}), 404

    return response




@api_blueprint.route('/get-report/<id>', methods=['GET'])
def get_report(id):
    temp_pdf_path = f"{PDF_DIR}/temp.pdf"
    output_pdf_path = f"{PDF_DIR}/final.pdf"
    try:
        try:
            organ_metrics = get_mask_data_internal(id)
            organ_metrics = organ_metrics.get("organ_metrics", [])
        except Exception as e:
            return jsonify({"error": f"Error loading organ metrics: {str(e)}"}), 500
 
        base_path = f"{SESSIONS_DIR}/{id}"
        # New flat structure, matching get-main-nifti / get-label-colormap above —
        # this fixes a bug from the merge where `subfolder`/`label_subfolder` were
        # referenced here but never defined, which would have raised a NameError
        # on every call to this route.
        ct_path = f"{Constants.PANTS_PATH}/image_only/{get_panTS_id(id)}/{Constants.MAIN_NIFTI_FILENAME}"
        masks = f"{Constants.PANTS_PATH}/mask_only/{get_panTS_id(id)}/{Constants.COMBINED_LABELS_NIFTI_FILENAME}"
 
        template_pdf = os.getenv("TEMPLATE_PATH", "report_template_3.pdf")
 
        extracted_data = None
        column_headers = None
        try:
            csv_path = f"{base_path}/info.csv"
            df = pd.read_csv(csv_path)
            extracted_data = df.iloc[0] if len(df) > 0 else None
            column_headers = df.columns.tolist()
        except Exception:
            pass
 
        generate_pdf_with_template(
            output_pdf=output_pdf_path,
            folder_name=id,
            ct_path=ct_path,
            mask_path=masks,
            template_pdf=template_pdf,
            temp_pdf_path=temp_pdf_path,
            id=id,
            extracted_data=extracted_data,
            column_headers=column_headers
        )
 
        return send_file(
            output_pdf_path,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"report_{id}.pdf"
        )
 
    except Exception as e:
        return jsonify({"error": f"Unhandled error: {str(e)}"}), 500
 
    finally:
        if os.path.exists(temp_pdf_path):
            os.remove(temp_pdf_path)
 
 
@api_blueprint.route('/explain-impressions', methods=['POST'])
def explain_impressions():
    """Deprecated: plain-language explanations required Ollama, which isn't
    available in production. The frontend no longer calls this button, but
    the route stays as a stable no-op so old clients don't 404."""
    return jsonify({
        "plain_language": ["Plain-language summary unavailable — please ask your doctor to walk through these findings with you."],
    }), 200
 
@api_blueprint.route('/define-term', methods=['GET'])
def define_term():
    """Returns a plain-English definition for a clicked medical term, from
    the hardcoded MEDICAL_TERM_DICTIONARY only — no LLM fallback."""
    term = (request.args.get('term') or '').strip().lower()
    if not term:
        return jsonify({"error": "Missing term parameter"}), 400

    if term in MEDICAL_TERM_DICTIONARY:
        return jsonify({"term": term, "definition": MEDICAL_TERM_DICTIONARY[term], "source": "dictionary"})

    return jsonify({
        "term": term,
        "definition": "Definition unavailable — try asking your doctor what this term means.",
        "source": "fallback",
    }), 200
 
 
@api_blueprint.route('/get-report-data/<id>', methods=['GET'])
def get_report_data(id):
    if id is None or not str(id).isdigit():
        return jsonify({"error": "Invalid id parameter"}), 400
    case_id = int(id)
    try:
        if id is None or not str(id).isdigit():
            return jsonify({"error": "Invalid id parameter"}), 400
        id = str(int(id))
        # ── Try RadGPT structured report from metadata.xlsx first ─────────────
        # This uses Zongwei Zhou's own RadGPT model output — more accurate
        # than Ollama-generated impressions. Falls back to Ollama if not found.
        radgpt_comments = None
        radgpt_impression = None
        try:
            import openpyxl, re
            metadata_path = os.path.join(Constants.PANTS_PATH, "data", "metadata.xlsx")
            if os.path.exists(metadata_path):
                wb = openpyxl.load_workbook(metadata_path, read_only=True, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows())]
                id_col = next((i for i, h in enumerate(headers) if h and 'ID' in str(h)), 0)
                report_col = next((i for i, h in enumerate(headers) if h and 'report' in str(h).lower()), -1)
                if report_col >= 0:
                    pants_id = get_panTS_id(id)
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_id = str(row[id_col] or '').strip()
                        if row_id == pants_id:
                            raw = str(row[report_col] or '')
                            # Clean Windows carriage return artifacts
                            raw = raw.replace('_x000D_', '\n').replace('\r\n', '\n').replace('\r', '\n')
                            # Collapse multiple blank lines
                            import re as _re
                            raw = _re.sub(r'\n{3,}', '\n\n', raw)
                            findings_match = re.search(r'FINDINGS:(.*?)(?=IMPRESSION:|$)', raw, re.DOTALL)
                            impression_match = re.search(r'IMPRESSION:(.*?)$', raw, re.DOTALL)
                            if findings_match:
                                radgpt_comments = findings_match.group(1).strip()
                            if impression_match:
                                imp_text = impression_match.group(1).strip()
                                # Keep full impression, split into sentences
                                sentences = [s.strip() for s in re.split(r'(?<=[.!?])\s+', imp_text) if s.strip()]
                                radgpt_impression = sentences if sentences else [imp_text]
                            print(f"[RadGPT] Found report for case {id}: {radgpt_impression}")
                            break
                wb.close()
        except Exception as e:
            print(f"[RadGPT] metadata lookup failed: {e}")
        # ─────────────────────────────────────────────────────────────────────
        subfolder = "ImageTr" if case_id < 9000 else "ImageTe"
        label_subfolder = "LabelTr" if case_id < 9000 else "LabelTe"
        # Check image_only first (new structure), fall back to data/ImageTr
        image_only_path = f"{Constants.PANTS_PATH}/image_only/{get_panTS_id(case_id)}/{Constants.MAIN_NIFTI_FILENAME}"
        data_ct_path = f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(case_id)}/{Constants.MAIN_NIFTI_FILENAME}"
        ct_path = image_only_path if os.path.exists(image_only_path) else data_ct_path
        # Check mask_only first (new structure), fall back to data/LabelTe
        mask_only_path = f"{Constants.PANTS_PATH}/mask_only/{get_panTS_id(case_id)}/{Constants.COMBINED_LABELS_NIFTI_FILENAME}"
        data_mask_path = f"{Constants.PANTS_PATH}/data/{label_subfolder}/{get_panTS_id(case_id)}/{Constants.COMBINED_LABELS_NIFTI_FILENAME}"
        mask_path = mask_only_path if os.path.exists(mask_only_path) else data_mask_path
        seg_dir = f"{Constants.PANTS_PATH}/data/{label_subfolder}/{get_panTS_id(case_id)}/segmentations"
 
        pid = get_panTS_id(case_id)
        meta = _METADATA_CACHE.get(pid, {})
        age = meta.get("age", "N/A")
        sex = meta.get("sex", "N/A")
 
        wb = load_workbook(os.path.join(Constants.PANTS_PATH, "data", "metadata.xlsx"))
        sheet = wb["PanTS_metadata"]
        contrast = ""
        study_detail = ""
        for row in sheet.iter_rows(values_only=True):
            if row[0] == pid:
                contrast = row[3]
                study_detail = row[8]
                break
 
        # If local files don't exist, download from HuggingFace
        if not os.path.exists(ct_path) or not os.path.exists(mask_path):
            import requests, tempfile
            pants_id = get_panTS_id(id)
            hf_base = f"https://huggingface.co/datasets/BodyMaps/iPanTSMini/resolve/main"
            tmp_dir = tempfile.mkdtemp()
            if not os.path.exists(ct_path):
                hf_ct = f"{hf_base}/image_only/{pants_id}/ct.nii.gz?download=true"
                ct_path = os.path.join(tmp_dir, "ct.nii.gz")
                print(f"[HuggingFace] Downloading CT for case {id}...")
                r = requests.get(hf_ct, stream=True, verify=False)
                with open(ct_path, 'wb') as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        f.write(chunk)
            if not os.path.exists(mask_path):
                hf_mask = f"{hf_base}/mask_only/{pants_id}/combined_labels.nii.gz?download=true"
                mask_path = os.path.join(tmp_dir, "combined_labels.nii.gz")
                print(f"[HuggingFace] Downloading mask for case {id}...")
                r = requests.get(hf_mask, stream=True, verify=False)
                with open(mask_path, 'wb') as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        f.write(chunk)
 
        ct_nii = nib.load(ct_path)
        spacing = ct_nii.header.get_zooms()
        shape = ct_nii.shape
        ct_array = ct_nii.get_fdata()
        mask_nii = nib.load(mask_path)
        mask_array = mask_nii.get_fdata().astype(np.uint8)
        # Crop both arrays to the minimum shape along each axis
        # to handle slight size mismatches between CT and mask
        min_shape = tuple(min(c, m) for c, m in zip(ct_array.shape, mask_array.shape))
        ct_array = ct_array[:min_shape[0], :min_shape[1], :min_shape[2]]
        mask_array = mask_array[:min_shape[0], :min_shape[1], :min_shape[2]]
        voxel_volume = np.prod(mask_nii.header.get_zooms()) / 1000
        # World-space affine - converts a voxel index (i, j, k) to real
        # millimeter coordinates. This is what makes the centroid below a
        # REAL position usable by moveCornerstoneCrosshairToMm, rather
        # than a placeholder.
        affine = mask_nii.affine
 
        LABELS = {v: k for k, v in Constants.PREDEFINED_LABELS.items()}
 
        # Soft physiological sanity ranges per organ type, used only to
        # flag "this needs review" vs "looks normal" internally. NOT
        # shown to the user as raw numbers - the frontend only ever sees
        # the `status` field. This is a stopgap for a known upstream
        # segmentation/data issue where some organs read in the air range;
        # rather than silently showing a wrong number as fact, or trying
        # to "correct" the data here, we flag it so the UI can say
        # "needs review" instead of presenting a confident wrong reading.
        SOLID_ORGAN_HU_RANGE = (-20, 150)      # liver, spleen, kidney, pancreas, etc.
        GI_HOLLOW_ORGAN_HU_RANGE = (-300, 200)  # colon, stomach, intestine, duodenum - tightened from
                                                 # -1000 so a mean this close to pure air (e.g. -756 for
                                                 # colon) correctly flags "check" instead of "normal" -
                                                 # a real colon has enough wall/stool tissue that a mean
                                                 # in the deep-air range usually signals a segmentation
                                                 # issue, not a genuinely normal reading.
        LUNG_HU_RANGE = (-1000, -200)            # lungs are genuinely air-filled - this range is correct as-is
        GI_HOLLOW_ORGANS = {"colon", "stomach", "intestine", "duodenum"}
        LUNG_ORGANS = {"lung_left", "lung_right"}
 
        organ_volumes = {}
        NO_FLAG_ORGANS = {
            "femur_left", "femur_right", "aorta", "postcava", "veins",
            "celiac_artery", "superior_mesenteric_artery", "renal_vein_left",
            "renal_vein_right", "common_bile_duct", "pancreatic_duct",
        }
        for organ, label_id in LABELS.items():
            if label_id == 0:
                continue
            mask = (mask_array == label_id)
            if not np.any(mask):
                continue
            volume = float(np.sum(mask) * voxel_volume)
            mean_hu = float(np.mean(ct_array[mask]))
 
            if organ in LUNG_ORGANS:
                lo, hi = LUNG_HU_RANGE
            elif organ in GI_HOLLOW_ORGANS:
                lo, hi = GI_HOLLOW_ORGAN_HU_RANGE
            else:
                lo, hi = SOLID_ORGAN_HU_RANGE
            status = "normal" if organ in NO_FLAG_ORGANS else ("check" if (mean_hu < lo or mean_hu > hi) else "normal")
 
            # Real centroid: voxel-space center of mass, converted to mm
            # via the affine. This is genuine anatomical position - not a
            # placeholder - and feeds the same crosshair-navigation
            # plumbing already used elsewhere (moveCornerstoneCrosshairToMm
            # / moveNiiVueCrosshairToMm) for click-to-jump.
            voxel_coords = np.argwhere(mask)
            centroid_voxel = voxel_coords.mean(axis=0)  # (i, j, k) in voxel space
            centroid_world = nib.affines.apply_affine(affine, centroid_voxel)
 
            # Bounding box dimensions in cm (real physical size of the organ)
            bbox_min = voxel_coords.min(axis=0)
            bbox_max = voxel_coords.max(axis=0)
            bbox_voxels = bbox_max - bbox_min + 1
            # Convert voxel counts to mm using spacing, then to cm
            spacing_mm = np.abs([affine[0,0], affine[1,1], affine[2,2]])
            dims_mm = bbox_voxels * spacing_mm
            dims_cm = [round(float(d)/10, 1) for d in dims_mm]
 
            organ_volumes[organ] = {
                "volume": round(volume, 2),
                "mean_hu": round(mean_hu, 1),
                "status": status,
                "centroid_mm": [round(float(c), 2) for c in centroid_world],
                "dimensions": dims_cm,
            }
 
        lesions = {}
        lesion_files = {
            "pancreas": "pancreatic_lesion.npz",
            "liver": "liver_lesion.npz",
            "kidney": "kidney_lesion.npz",
        }
        for organ, filename in lesion_files.items():
            path = os.path.join(seg_dir, filename)
            if os.path.exists(path):
                data = np.load(path)["data"]
                voxels = int(np.sum(data > 0))
                if voxels > 0:
                    lesion_volume = round(voxels * voxel_volume, 2)
                    lesions[organ] = {"voxels": voxels, "volume": lesion_volume}
 
        organ_data_str = ""
        for organ, vals in organ_volumes.items():
            organ_data_str += f"{organ.replace('_', ' ')}: volume={vals['volume']}cc, mean HU={vals['mean_hu']}\n"
 
        # If we have a RadGPT report, it is the authoritative source for organ status.
        # Reset everything to normal first, then flag only what RadGPT calls abnormal.
        if radgpt_comments:
            for organ in list(organ_volumes.keys()):
                organ_volumes[organ]['status'] = 'normal'
            abnormal_keywords = ['enlarged', 'mass', 'lesion', 'tumor', 'abnormal',
                                 'dilated', 'obstruction', 'isoattenuating', 'hypodense',
                                 'hyperdense', 'cyst', 'nodule', 'atrophy', 'bilateral']
            # Build stripped root for flexible matching
            organ_roots = {}
            for organ in organ_volumes.keys():
                root = organ.replace('_left','').replace('_right','').replace('_body','') \
                            .replace('_head','').replace('_tail','').replace('_gland','') \
                            .replace('_duct','').replace('_lesion','').replace('_','')
                organ_roots[organ] = root.lower()
            for line in radgpt_comments.split('\n'):
                line_stripped = line.lower().replace(' ','').replace('_','')
                if any(kw in line.lower() for kw in abnormal_keywords):
                    for organ, root in organ_roots.items():
                        # Skip subtypes unless explicitly mentioned
                        # e.g. "pancreas enlarged" shouldn't flag pancreas_body/head/tail
                        if '_body' in organ or '_head' in organ or '_tail' in organ or '_duct' in organ:
                            # Only flag subtype if the subtype word is in the line
                            subtype = organ.split('_')[-1]
                            if root in line_stripped and subtype in line.lower():
                                organ_volumes[organ]['status'] = 'check'
                        else:
                            if root in line_stripped:
                                organ_volumes[organ]['status'] = 'check'
        comments = radgpt_comments or "Clinical comments unavailable."
        impression_items = radgpt_impression or ["No impression available for this case."]
 
        return jsonify({
            "case_id": id,
            "patient": {"age": age, "sex": sex},
            "imaging": {
                "study_type": study_detail,
                "contrast": contrast,
                "spacing": [round(float(s), 3) for s in spacing],
                "shape": list(shape),
            },
            "organ_volumes": organ_volumes,
            "lesions": lesions,
            "comments": comments,
            "impression": impression_items,
        })
 
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": "An internal error occurred."}), 500
 
 
@api_blueprint.route('/get-specific-segmentations/<combined_labels_id>', methods=['POST'])
async def get_specific_segmentations(combined_labels_id):
    combined_labels_id = combined_labels_id.replace("PanTS_", "")
    combined_labels_id = combined_labels_id.lstrip("0")
    try: 
        organs = json.loads(request.form["organs"])
        niftiProcessor = NpzProcessor()
        combined_labels, intensities, affine, header = niftiProcessor.nifti_combine_labels(int(combined_labels_id), {"foo": "bar"}, save=False, organs=organs)
        if combined_labels.dtype != np.uint8:
            data_uint8 = combined_labels.astype(np.uint8)
            new_img = nib.Nifti1Image(data_uint8, affine, header=header)
            new_img.set_data_dtype(np.uint8)
        else:
            new_img = nib.Nifti1Image(combined_labels, affine, header=header)
            new_img.set_data_dtype(np.uint8)
        
        buffer = io.BytesIO()
        file_map = new_img.make_file_map()
        file_map['image'].fileobj = buffer
        new_img.to_file_map(file_map)

        buffer.seek(0)
        
        response = send_file(buffer, as_attachment=True, download_name="combined_specific_labels.nii.gz", mimetype="application/gzip")
        response.headers["Cross-Origin-Resource-Policy"] = "cross-origin"
        return response
    except Exception as e: 
        return jsonify({"error": f"Error loading organ metrics: {str(e)}"}), 500
@api_blueprint.route('/get-segmentations/<combined_labels_id>.nii.gz', methods=['GET'])
async def get_segmentations(combined_labels_id):
    if not _is_safe_id(combined_labels_id):
        return jsonify({"error": "Invalid id"}), 400
    nifti_path = f"{Constants.PANTS_PATH}/mask_only/{get_panTS_id(secure_filename(combined_labels_id))}/{Constants.COMBINED_LABELS_NIFTI_FILENAME}"
    labels = list(Constants.PREDEFINED_LABELS.values())
    # ?res=low → serve the precomputed low-res mask (paired with the low-res CT so the
    # overlay stays aligned). Falls back to full res below if it hasn't been generated.
    if (request.args.get('res') or '').strip().lower() == 'low':
        low_path = nifti_path.replace('.nii.gz', '_lowres.nii.gz')
        if os.path.exists(low_path):
            response = make_response(send_file(low_path, mimetype='application/gzip'))
            response.headers["Cross-Origin-Resource-Policy"] = "cross-origin"
            response.headers['Cache-Control'] = 'public, max-age=604800, immutable'
            return response

    img = nib.load(nifti_path)

    try:
        serve_path = nifti_path
        if img.get_data_dtype() != np.uint8:
            # The source is a float label map; Cornerstone needs uint8. Cache the
            # converted copy in a WRITABLE temp dir and serve THAT — never write
            # into the dataset's mask_only/ (it is read-only on the server, and an
            # HTTP GET must not mutate ground-truth data). Writing the sibling into
            # mask_only/ 500'd the segmentation endpoint in production.
            cache_dir = "/tmp/pants_uint8"
            os.makedirs(cache_dir, exist_ok=True)
            converted_path = os.path.join(
                cache_dir,
                f"{get_panTS_id(secure_filename(combined_labels_id))}_combined_labels_uint8.nii.gz",
            )
            if not os.path.exists(converted_path):
                print("⚠️ Detected float label map, converting to uint8 for Cornerstone compatibility...")
                raw = np.asanyarray(img.dataobj)
                data = np.rint(raw).astype(np.uint8)

                new_img = nib.Nifti1Image(data, img.affine, header=img.header)
                new_img.set_data_dtype(np.uint8)
                nib.save(new_img, converted_path)
            serve_path = converted_path

        response = make_response(send_file(serve_path, mimetype='application/gzip'))
        response.headers["Cross-Origin-Resource-Policy"] = "cross-origin"
        response.headers['Cache-Control'] = 'public, max-age=604800, immutable'
        # response.headers['Cross-Origin-Opener-Policy'] = 'same-origin'
        # response.headers['Cross-Origin-Embedder-Policy'] = 'require-corp'

        return response

    except Exception as e:
        print(f"❌ [get-segmentations ERROR] {e}")
        return jsonify({"error": str(e)}), 500


@api_blueprint.route('/download/<id>', methods=['GET'])
def download_segmentation_zip(id):
    try:
        if not _is_safe_id(id):
            return jsonify({"error": "Invalid id"}), 400
        outputs_ct_folder = Path(f"{Constants.PANTS_PATH}/mask_only/{get_panTS_id(secure_filename(str(id)))}/segmentations")
        
        if not os.path.exists(outputs_ct_folder):
            return jsonify({"error": "Outputs/ct folder not found"}), 404
        
        files = list(outputs_ct_folder.glob("*"))

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for file_path in files:
                zip_file.write(file_path, arcname=file_path.name) 

        zip_buffer.seek(0)  # rewind

        return send_file(
            zip_buffer,
            mimetype="application/zip",
            as_attachment=True,
            download_name=f"case_{id}_segmentations.zip"
        )



    except Exception as e:
        print(f"❌ [Download Error] {e}")
        return jsonify({"error": "Internal server error"}), 500

import time

inference_jobs = {}  # {session_id: {status, model, error, session_path, zip_path}}
# Guards the read-modify-write in _set_inference_job: background segmentation
# threads and request handlers touch inference_jobs concurrently, and the
# get/update/set below is not atomic without a lock (updates would be lost).
_inference_jobs_lock = threading.Lock()


def _set_inference_job(session_id, **kwargs):
    with _inference_jobs_lock:
        current = inference_jobs.get(session_id, {})
        current.update(kwargs)
        inference_jobs[session_id] = current


def _start_auto_segmentation(session_id, model_name, ct_file=None, server_input_path=None):
    if not _is_safe_id(session_id):
        return jsonify({"error": "Invalid session ID"}), 400
    session_path = os.path.join(SESSIONS_DIR, session_id)
    os.makedirs(session_path, exist_ok=True)

    if model_name == 'ShapeKit':
        # ShapeKit takes a segmentation directory from a previous step, not a CT file
        if not server_input_path or not os.path.isdir(server_input_path):
            return jsonify({"error": "ShapeKit requires INPUT_SERVER_PATH pointing to a segmentation output directory"}), 400
        input_path = server_input_path
    elif ct_file is not None:
        input_path = os.path.join(session_path, ct_file.filename)
        ct_file.save(input_path)
    elif server_input_path:
        if not os.path.exists(server_input_path):
            return jsonify({"error": f"INPUT_SERVER_PATH does not exist: {server_input_path}"}), 400
        input_path = os.path.join(session_path, os.path.basename(server_input_path))
        import shutil
        shutil.copy2(server_input_path, input_path)
    else:
        return jsonify({"error": "No CT file provided. Send MAIN_NIFTI or INPUT_SERVER_PATH."}), 400

    _set_inference_job(
        session_id,
        status="running",
        model=model_name,
        error=None,
        ct_path=input_path,
        session_path=session_path,
        zip_path=os.path.join(session_path, "auto_masks.zip"),
    )

    def do_segmentation_and_zip():
        try:
            output_mask_dir = run_auto_segmentation(input_path, session_dir=session_path, model=model_name)

            if output_mask_dir is None or not os.path.exists(output_mask_dir):
                msg = f"Auto segmentation failed for session {session_id}"
                print(f"❌ {msg}")
                _set_inference_job(session_id, status="failed", error=msg)
                return

            zip_path = os.path.join(session_path, "auto_masks.zip")
            with zipfile.ZipFile(zip_path, 'w') as zipf:
                for dirpath, _, filenames in os.walk(output_mask_dir):
                    for filename in filenames:
                        if filename.endswith(".nii.gz") or (model_name == "ePAI" and filename.endswith(".csv")):
                            abs_path = os.path.join(dirpath, filename)
                            arcname = os.path.relpath(abs_path, output_mask_dir)
                            zipf.write(abs_path, arcname=arcname)

            if session_id in progress_tracker:
                start_time, expected_time, _ = progress_tracker[session_id]
                progress_tracker[session_id] = (start_time, expected_time, True)
                progress_tracker.pop(session_id, None)

            _set_inference_job(session_id, status="completed", error=None,
                               zip_path=zip_path, output_mask_dir=output_mask_dir)
            print(f"✅ Finished segmentation and zipping for session {session_id}")
        except Exception as e:
            print(f"❌ Exception while processing session {session_id}: {e}")
            _set_inference_job(session_id, status="failed", error=str(e))

    threading.Thread(target=do_segmentation_and_zip, daemon=True).start()
    print("[Server] auto_segment request is returning now")
    return jsonify({"message": "Segmentation started", "session_id": session_id}), 200

@api_blueprint.route('/auto_segment/<session_id>', methods=['POST'])
def auto_segment(session_id):

    model_name = request.form.get("MODEL_NAME", None)
    server_input_path = request.form.get("INPUT_SERVER_PATH", None)

    ct_file = request.files.get('MAIN_NIFTI')

    # Check if model name is valid
    if model_name is None:
        return {"error": "MODEL_NAME is required."}, 400
    return _start_auto_segmentation(
        session_id=session_id,
        model_name=model_name,
        ct_file=ct_file,
        server_input_path=server_input_path,
    )


@api_blueprint.route('/run-epai-inference', methods=['POST'])
@api_blueprint.route('/run-inference', methods=['POST'])
def run_epai_inference():
    """
    Runs ePAI inference with either:
      1) multipart file: MAIN_NIFTI
      2) server path: INPUT_SERVER_PATH

    Optional fields:
      - session_id
      - uploaded_filename (used with chunked upload output in sessions/inference/<session_id>/)
    """
    payload = request.get_json(silent=True) or {}

    def _pick_text(*keys):
        for key in keys:
            value = request.form.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
        for key in keys:
            value = payload.get(key) if isinstance(payload, dict) else None
            if isinstance(value, str) and value.strip():
                return value.strip()
        return None

    session_id = _pick_text("session_id", "SESSION_ID", "sessionId") or str(uuid.uuid4())
    model_name = _pick_text("model_name", "model", "MODEL_NAME") or "ePAI"
    uploaded_filename = _pick_text("uploaded_filename", "output_filename", "filename")
    input_server_path = _pick_text("INPUT_SERVER_PATH", "input_server_path", "server_path", "path")
    source_reconstruction_session_id = _pick_text("source_reconstruction_session_id")
    ct_file = (
        request.files.get('MAIN_NIFTI')
        or request.files.get('file')
        or request.files.get('ct')
        or request.files.get('ct_file')
    )

    if source_reconstruction_session_id and not input_server_path and ct_file is None:
        source_job = inference_jobs.get(source_reconstruction_session_id, {})
        source_output_dir = source_job.get("output_mask_dir")
        if source_output_dir:
            recon_path = os.path.join(source_output_dir, "reconstructed_ct.nii.gz")
            if os.path.exists(recon_path):
                input_server_path = recon_path
            else:
                return jsonify({"error": f"Reconstructed CT not found for session {source_reconstruction_session_id}"}), 404
        else:
            return jsonify({"error": f"Source reconstruction session {source_reconstruction_session_id} not found or not completed"}), 404

    if not input_server_path and uploaded_filename:
        candidate = os.path.join(Constants.SESSIONS_DIR_NAME, "inference", session_id, uploaded_filename)
        if os.path.exists(candidate):
            input_server_path = candidate

    if not input_server_path and ct_file is None:
        inference_root = os.path.join(Constants.SESSIONS_DIR_NAME, "inference")
        infer_dir = os.path.join(inference_root, session_id)

        if not os.path.isdir(infer_dir) and os.path.isdir(inference_root):
            candidate_dirs = []
            for dirname in os.listdir(inference_root):
                if dirname == session_id or dirname.startswith(session_id) or session_id.startswith(dirname):
                    full_dir = os.path.join(inference_root, dirname)
                    if os.path.isdir(full_dir):
                        candidate_dirs.append(full_dir)
            if len(candidate_dirs) == 1:
                infer_dir = candidate_dirs[0]

        if os.path.isdir(infer_dir):
            nii_candidates = []
            for root, _, files in os.walk(infer_dir):
                for name in files:
                    lower = name.lower()
                    if lower.endswith(".nii.gz") or lower.endswith(".nii"):
                        nii_candidates.append(os.path.join(root, name))

            if nii_candidates:
                preferred = [p for p in nii_candidates if os.path.basename(p).lower() in {"ct.nii.gz", "ct.nii"}]
                pool = preferred if preferred else nii_candidates
                pool.sort(key=lambda p: os.path.getmtime(p), reverse=True)
                input_server_path = pool[0]

    return _start_auto_segmentation(
        session_id=session_id,
        model_name=model_name,
        ct_file=ct_file,
        server_input_path=input_server_path,
    )


@api_blueprint.route('/inference-status/<session_id>', methods=['GET'])
def get_inference_status(session_id):
    job = inference_jobs.get(session_id)
    if job is None:
        return jsonify({"status": "not_found", "session_id": session_id}), 404
    return jsonify({"session_id": session_id, **job}), 200


@api_blueprint.route('/check-inference-status', methods=['GET'])
def check_inference_status_legacy():
    session_id = request.args.get("session_id") or request.args.get("sessionId")
    if not session_id:
        return jsonify({"error": "session_id is required"}), 400
    return get_inference_status(session_id)


@api_blueprint.route('/jobs', methods=['POST'])
def create_pull_inference_job():
    payload = request.get_json(silent=True) or {}

    def _pick_text(*keys):
        for key in keys:
            value = request.form.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
        for key in keys:
            value = payload.get(key) if isinstance(payload, dict) else None
            if isinstance(value, str) and value.strip():
                return value.strip()
        return None

    session_id = _pick_text("session_id", "SESSION_ID", "sessionId") or str(uuid.uuid4())
    input_server_path = _pick_text("input_server_path", "INPUT_SERVER_PATH", "path")
    uploaded_filename = _pick_text("uploaded_filename", "output_filename", "filename")
    model = _pick_text("model", "MODEL_NAME") or "ePAI"

    ct_file = (
        request.files.get('MAIN_NIFTI')
        or request.files.get('file')
        or request.files.get('ct')
        or request.files.get('ct_file')
    )

    if not input_server_path and uploaded_filename:
        candidate = os.path.join(Constants.SESSIONS_DIR_NAME, "inference", session_id, uploaded_filename)
        if os.path.exists(candidate):
            input_server_path = candidate

    if ct_file is not None and not input_server_path:
        target_dir = os.path.join(Constants.SESSIONS_DIR_NAME, "inference", session_id, "input")
        os.makedirs(target_dir, exist_ok=True)
        original_name = os.path.basename(ct_file.filename or "ct.nii.gz")
        target_name = original_name if original_name else "ct.nii.gz"
        input_server_path = os.path.join(target_dir, target_name)
        ct_file.save(input_server_path)

    if not input_server_path:
        return jsonify({"error": "No input provided. Send input_server_path, uploaded_filename, or MAIN_NIFTI."}), 400
    if not os.path.exists(input_server_path):
        return jsonify({"error": f"Input path not found: {input_server_path}"}), 400

    job = inference_job_queue.create_job(
        input_file_path=input_server_path,
        session_id=session_id,
        model=model,
        max_attempts=int(os.getenv("INFERENCE_MAX_ATTEMPTS", "3")),
    )
    return jsonify(_public_job_payload(job)), 201


@api_blueprint.route('/jobs/<job_id>', methods=['GET'])
def get_pull_inference_job(job_id):
    job = inference_job_queue.get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found", "job_id": job_id}), 404
    return jsonify(_public_job_payload(job)), 200


@api_blueprint.route('/jobs/next', methods=['GET'])
def lease_next_pull_job():
    auth_error = _require_worker_auth()
    if auth_error is not None:
        return auth_error

    worker_id = _get_worker_id()
    lease_seconds = int(request.args.get("lease_seconds") or os.getenv("INFERENCE_LEASE_SECONDS", "900"))
    job = inference_job_queue.lease_next_job(worker_id=worker_id, lease_seconds=lease_seconds)
    if not job:
        return ("", 204)

    job_id = job.get("job_id")
    return jsonify({
        "job_id": job_id,
        "session_id": job.get("session_id"),
        "model": job.get("model"),
        "lease_seconds": lease_seconds,
        "input_download_url": _absolute_api_url(f"/jobs/{job_id}/input"),
        "heartbeat_url": _absolute_api_url(f"/jobs/{job_id}/heartbeat"),
        "result_upload_url": _absolute_api_url(f"/jobs/{job_id}/result"),
        "fail_url": _absolute_api_url(f"/jobs/{job_id}/fail"),
        "status_url": _absolute_api_url(f"/jobs/{job_id}"),
    }), 200


@api_blueprint.route('/jobs/<job_id>/input', methods=['GET'])
def download_pull_job_input(job_id):
    auth_error = _require_worker_auth()
    if auth_error is not None:
        return auth_error

    job = inference_job_queue.get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found", "job_id": job_id}), 404

    input_path = job.get("input_file_path")
    if not input_path or not os.path.exists(input_path):
        return jsonify({"error": "Input file missing for job", "job_id": job_id}), 404

    return send_file(input_path, as_attachment=True, download_name=os.path.basename(input_path))


@api_blueprint.route('/jobs/<job_id>/heartbeat', methods=['POST'])
def heartbeat_pull_job(job_id):
    auth_error = _require_worker_auth()
    if auth_error is not None:
        return auth_error

    worker_id = _get_worker_id()
    payload = request.get_json(silent=True) or {}
    lease_seconds = int(payload.get("lease_seconds") or request.form.get("lease_seconds") or os.getenv("INFERENCE_LEASE_SECONDS", "900"))

    try:
        job = inference_job_queue.heartbeat(job_id=job_id, worker_id=worker_id, lease_seconds=lease_seconds)
        if not job:
            return jsonify({"error": "Job not found", "job_id": job_id}), 404
        return jsonify(_public_job_payload(job)), 200
    except PermissionError as e:
        return jsonify({"error": str(e)}), 403
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@api_blueprint.route('/jobs/<job_id>/result', methods=['POST'])
def complete_pull_job(job_id):
    auth_error = _require_worker_auth()
    if auth_error is not None:
        return auth_error

    worker_id = _get_worker_id()
    prediction_file = (
        request.files.get("prediction")
        or request.files.get("mask")
        or request.files.get("combined_labels")
        or request.files.get("file")
    )
    output_csv = request.files.get("output_csv") or request.files.get("csv")

    if prediction_file is None:
        return jsonify({"error": "Missing prediction file (use field: prediction)"}), 400

    temp_dir = os.path.join("/tmp", "pull_job_results", job_id)
    os.makedirs(temp_dir, exist_ok=True)

    pred_path = os.path.join(temp_dir, "combined_labels.nii.gz")
    prediction_file.save(pred_path)

    csv_path = None
    if output_csv is not None:
        csv_path = os.path.join(temp_dir, "output.csv")
        output_csv.save(csv_path)

    try:
        job = inference_job_queue.complete_job(
            job_id=job_id,
            worker_id=worker_id,
            result_mask_path=pred_path,
            result_csv_path=csv_path,
        )
        if not job:
            return jsonify({"error": "Job not found", "job_id": job_id}), 404
        return jsonify(_public_job_payload(job)), 200
    except PermissionError as e:
        return jsonify({"error": str(e)}), 403
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 400


@api_blueprint.route('/jobs/<job_id>/fail', methods=['POST'])
def fail_pull_job(job_id):
    auth_error = _require_worker_auth()
    if auth_error is not None:
        return auth_error

    worker_id = _get_worker_id()
    payload = request.get_json(silent=True) or {}
    error_message = payload.get("error") or request.form.get("error") or "Worker marked job failed"

    try:
        job = inference_job_queue.fail_job(job_id=job_id, worker_id=worker_id, error=error_message)
        if not job:
            return jsonify({"error": "Job not found", "job_id": job_id}), 404
        return jsonify(_public_job_payload(job)), 200
    except PermissionError as e:
        return jsonify({"error": str(e)}), 403


@api_blueprint.route('/jobs/<job_id>/download', methods=['GET'])
def download_pull_job_result(job_id):
    job = inference_job_queue.get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found", "job_id": job_id}), 404
    if job.get("status") != "succeeded":
        return jsonify({"error": "Result not ready", "status": job.get("status")}), 409

    zip_path = job.get("result_zip_path")
    if not zip_path or not os.path.exists(zip_path):
        return jsonify({"error": "Result archive missing"}), 404

    return send_file(zip_path, as_attachment=True, download_name=f"{job_id}_auto_masks.zip")



@api_blueprint.route('/get_result/<session_id>', methods=['GET'])
def get_result(session_id):
    if not _is_safe_id(session_id):
        return jsonify({"error": "Invalid session ID"}), 400
    session_path = os.path.join(SESSIONS_DIR, session_id)
    zip_path = os.path.join(session_path, "auto_masks.zip")

    # Poll briefly for the archive. If it isn't ready, return 202 so the client
    # can keep polling instead of receiving an uncaught TimeoutError → 500.
    try:
        wait_for_file(zip_path, timeout=30)
    except TimeoutError:
        return jsonify({"error": "Result not ready", "session_id": session_id}), 202

    response = send_file(
        zip_path,
        as_attachment=True,
        download_name="auto_masks.zip"
    )
    response.headers["X-Session-Id"] = session_id
    return response


@api_blueprint.route('/session-ct/<session_id>', methods=['GET'])
def get_session_ct(session_id):
    job = inference_jobs.get(session_id, {})
    ct_path = job.get("ct_path")
    if not ct_path or not os.path.exists(ct_path):
        return jsonify({"error": "CT file not found for session"}), 404
    response = make_response(send_file(ct_path, mimetype='application/gzip'))
    response.headers['Cross-Origin-Resource-Policy'] = 'cross-origin'
    response.headers['Content-Encoding'] = 'gzip'
    return response


@api_blueprint.route('/session-segmentation/<session_id>', methods=['GET'])
def get_session_segmentation(session_id):
    job = inference_jobs.get(session_id, {})
    output_mask_dir = job.get("output_mask_dir")
    if not output_mask_dir:
        return jsonify({"error": "Segmentation not ready for session"}), 404
    seg_path = os.path.join(output_mask_dir, "combined_labels.nii.gz")
    if not os.path.exists(seg_path):
        return jsonify({"error": f"combined_labels.nii.gz not found at {seg_path}"}), 404
    response = make_response(send_file(seg_path, mimetype='application/gzip'))
    response.headers['Cross-Origin-Resource-Policy'] = 'cross-origin'
    response.headers['Content-Encoding'] = 'gzip'
    return response


@api_blueprint.route('/session-reconstruction/<session_id>', methods=['GET'])
def get_session_reconstruction(session_id):
    """Serves the OpenVAE reconstructed CT for a session."""
    job = inference_jobs.get(session_id, {})
    output_mask_dir = job.get("output_mask_dir")
    if not output_mask_dir:
        return jsonify({"error": "Reconstruction not ready for session"}), 404
    recon_path = os.path.join(output_mask_dir, "reconstructed_ct.nii.gz")
    if not os.path.exists(recon_path):
        return jsonify({"error": f"reconstructed_ct.nii.gz not found"}), 404
    response = make_response(send_file(recon_path, mimetype='application/gzip'))
    response.headers['Cross-Origin-Resource-Policy'] = 'cross-origin'
    response.headers['Content-Encoding'] = 'gzip'
    return response

#
#@api_blueprint.route('/progress_end/<session_id>', methods=['GET'])
#def progress_end(session_id):
#    progress_tracker.pop(session_id, None)
#    return jsonify({"message": "Progress End"}), 200


#### INFERENCE ENDPOINTS ####

CHUNK_DIR = "/tmp/uploads"  # Temporary folder for chunked uploads
os.makedirs(CHUNK_DIR, exist_ok=True)

@api_blueprint.route("/upload-inference-chunk", methods=["POST"])
def upload_inference_chunk():
    """
    Receives a chunk of a file.
    Expects:
        - session_id
        - chunk_index
        - total_chunks
        - file (the chunk itself)
    """
    try:
        session_id = request.form.get("session_id")
        chunk_index = request.form.get("chunk_index")
        total_chunks = request.form.get("total_chunks")
        chunk_file = request.files.get("file")

        if not all([session_id, chunk_index, total_chunks, chunk_file]):
            return jsonify({"error": "Missing parameters"}), 400

        # session_id and chunk_index are both joined into a filesystem path below;
        # reject anything non-numeric / traversal-y before it touches os.path.
        if not _is_safe_id(session_id):
            return jsonify({"error": "Invalid session ID"}), 400
        if not str(chunk_index).isdigit():
            return jsonify({"error": "Invalid chunk index"}), 400

        session_folder = os.path.join(CHUNK_DIR, session_id)
        os.makedirs(session_folder, exist_ok=True)

        chunk_path = os.path.join(session_folder, f"chunk-{chunk_index}")
        chunk_file.save(chunk_path)

        return jsonify({"status": "ok", "chunk_index": chunk_index})
    except Exception as e:
        print(f"❌ Chunk upload error: {e}")
        return jsonify({"error": str(e)}), 500


@api_blueprint.route("/finalize-upload", methods=["POST"])
def finalize_upload():
    """
    Combines all chunks into the final file in the proper sessions folder.
    Expects:
        - session_id
        - total_chunks
        - output_filename (optional)
    """
    try:
        session_id = request.form.get("session_id")
        if not _is_safe_id(session_id):
            return jsonify({"error": "Invalid session ID"}), 400
        total_chunks = int(request.form.get("total_chunks"))
        output_filename = request.form.get("output_filename", "inference_input.gz")
        requested_bdmap_id = request.form.get("bdmap_id") or request.form.get("case_id")

        if requested_bdmap_id and requested_bdmap_id.strip():
            bdmap_id = requested_bdmap_id.strip()
            if not bdmap_id.startswith("BDMAP_"):
                bdmap_id = f"BDMAP_{bdmap_id}"
        else:
            digits = "".join(ch for ch in (session_id or "") if ch.isdigit())
            if len(digits) < 8:
                fallback_digits = f"{(uuid.uuid5(uuid.NAMESPACE_DNS, session_id or str(uuid.uuid4())).int % (10 ** 8)):08d}"
                digits = (digits + fallback_digits)[:8]
            else:
                digits = digits[:8]
            bdmap_id = f"BDMAP_{digits}"

        # New base path: sessions/inference/<session_id>/
        base_path = os.path.join(Constants.SESSIONS_DIR_NAME, "inference", session_id)
        os.makedirs(base_path, exist_ok=True)

        normalized_output = output_filename.strip()
        lower_name = normalized_output.lower()
        if lower_name.endswith(".nii.gz"):
            target_filename = "ct.nii.gz"
        elif lower_name.endswith(".nii"):
            target_filename = "ct.nii"
        else:
            target_filename = normalized_output

        target_dir = base_path
        if target_filename.lower().endswith(".nii") or target_filename.lower().endswith(".nii.gz"):
            target_dir = os.path.join(base_path, bdmap_id)
            os.makedirs(target_dir, exist_ok=True)

        final_path = os.path.join(target_dir, target_filename)

        # Combine chunks
        temp_folder = os.path.join("/tmp/uploads", session_id)
        with open(final_path, "wb") as out_file:
            for i in range(total_chunks):
                chunk_path = os.path.join(temp_folder, f"chunk-{i}")
                with open(chunk_path, "rb") as f:
                    out_file.write(f.read())

        # Optional: clean up temp chunks
        for i in range(total_chunks):
            os.remove(os.path.join(temp_folder, f"chunk-{i}"))
        os.rmdir(temp_folder)

        uploaded_filename = os.path.relpath(final_path, base_path)
        return jsonify({
            "status": "combined",
            "path": final_path,
            "bdmap_id": bdmap_id,
            "uploaded_filename": uploaded_filename,
        })
    except Exception as e:
        print(f"❌ Finalize upload error: {e}")
        return jsonify({"error": str(e)}), 500

@api_blueprint.route("/upload-dicom-slice", methods=["POST"])
def upload_dicom_slice():
    """Save a single DICOM slice to a session-specific temp directory."""
    try:
        session_id = request.form.get("session_id")
        if not session_id:
            return jsonify({"error": "session_id required"}), 400
        slice_file = request.files.get("file")
        if not slice_file:
            return jsonify({"error": "file required"}), 400

        dicom_dir = os.path.join("/tmp/uploads", session_id, "dicom")
        os.makedirs(dicom_dir, exist_ok=True)
        save_path = os.path.join(dicom_dir, slice_file.filename or f"{uuid.uuid4()}.dcm")
        slice_file.save(save_path)
        return jsonify({"status": "ok", "filename": os.path.basename(save_path)})
    except Exception as e:
        print(f"❌ DICOM slice upload error: {e}")
        return jsonify({"error": str(e)}), 500


@api_blueprint.route("/finalize-dicom", methods=["POST"])
def finalize_dicom():
    """Convert an uploaded DICOM series to NIfTI using SimpleITK."""
    try:
        import SimpleITK as sitk

        session_id = request.form.get("session_id")
        if not session_id:
            return jsonify({"error": "session_id required"}), 400

        dicom_dir = os.path.join("/tmp/uploads", session_id, "dicom")
        if not os.path.isdir(dicom_dir):
            return jsonify({"error": "No DICOM slices found for this session"}), 400

        reader = sitk.ImageSeriesReader()
        series_ids = reader.GetGDCMSeriesIDs(dicom_dir)
        if not series_ids:
            return jsonify({"error": "No valid DICOM series found in uploaded files"}), 400

        dicom_names = reader.GetGDCMSeriesFileNames(dicom_dir, series_ids[0])
        reader.SetFileNames(dicom_names)
        image = reader.Execute()
        image = sitk.DICOMOrient(image, "LPS")

        # Save to sessions/inference/<session_id>/<bdmap_id>/ct.nii.gz
        digits = "".join(ch for ch in (session_id or "") if ch.isdigit())
        if len(digits) < 8:
            fallback = f"{(uuid.uuid5(uuid.NAMESPACE_DNS, session_id).int % (10 ** 8)):08d}"
            digits = (digits + fallback)[:8]
        else:
            digits = digits[:8]
        bdmap_id = f"BDMAP_{digits}"

        base_path = os.path.join(Constants.SESSIONS_DIR_NAME, "inference", session_id)
        target_dir = os.path.join(base_path, bdmap_id)
        os.makedirs(target_dir, exist_ok=True)
        final_path = os.path.join(target_dir, "ct.nii.gz")

        sitk.WriteImage(image, final_path)

        # Clean up temp DICOM slices
        import shutil
        shutil.rmtree(os.path.join("/tmp/uploads", session_id), ignore_errors=True)

        uploaded_filename = os.path.relpath(final_path, base_path)
        return jsonify({
            "status": "converted",
            "path": final_path,
            "bdmap_id": bdmap_id,
            "uploaded_filename": uploaded_filename,
        })
    except Exception as e:
        print(f"❌ DICOM finalize error: {e}")
        return jsonify({"error": str(e)}), 500


## OTHER ENDPOINTS ##

@api_blueprint.route('/cancel-inference', methods=['POST'])
def cancel_inference():
    cancel_all_inference()
    for session_id, job in inference_jobs.items():
        if job.get('status') == 'running':
            _set_inference_job(session_id, status='failed', error='Cancelled by user')
    return jsonify({"message": "Inference cancelled"}), 200


@api_blueprint.route('/ping', methods=['GET'])
def ping():
    return jsonify({"message": "pong"}), 200

@api_blueprint.route("/search", methods=["GET"])
def api_search():
    # return jsonify({"message": "pong"}), 200
    df = apply_filters(DF).copy()
    df = ensure_sort_cols(df)

    # ---- 排序參數 ----
    sort_by  = (_arg("sort_by", "top") or "top").strip().lower()
    sort_dir = (_arg("sort_dir", "asc") or "asc").strip().lower()

    if sort_by in ("top", "quality"):
        by  = ["__complete", "__spacing_sum", "__shape_sum", "__case_sortkey"]
        asc = [False, True, False, True]
    elif sort_by in ("id", "id_asc"):
        by, asc = ["__case_sortkey"], [True]
    elif sort_by == "id_desc":
        by, asc = ["__case_sortkey"], [False]
    elif sort_by in ("shape_desc", "shape"):
        by, asc = ["__shape_sum", "__case_sortkey"], [False, True]
    elif sort_by in ("spacing_asc", "spacing"):
        by, asc = ["__spacing_sum", "__case_sortkey"], [True, True]
    elif sort_by == "age_asc":
        by, asc = ["__age", "__case_sortkey"], [True, True]
    elif sort_by == "age_desc":
        by, asc = ["__age", "__case_sortkey"], [False, True]
    else:
        key_map = {"id": "__case_sortkey", "spacing": "__spacing_sum", "shape": "__shape_sum"}
        k = key_map.get(sort_by, "__case_sortkey")
        by, asc = [k, "__case_sortkey"], [(sort_dir != "desc"), True]

    # ---- 排序 ----
    df = df.sort_values(by=by, ascending=asc, na_position="last", kind="mergesort")

    # ---- 分頁：注意 total 先算完篩選後的完整筆數 ----
    total    = int(len(df))
    page     = max(to_int(_arg("page", "1")) or 1, 1)
    per_page = to_int(_arg("per_page", "24")) or 24
    per_page = max(1, min(per_page, 1_000_000))

    pages = max(1, int(math.ceil(total / per_page)))
    page  = max(1, min(page, pages))
    start, end = (page - 1) * per_page, (page - 1) * per_page + per_page

    # ---- 轉成前端想要的 items ----
    items = [row_to_item(r) for _, r in df.iloc[start:end].iterrows()]
    items = clean_json_list(items)

    return jsonify({
        "items": items,         # ← 前端只讀這個渲染卡片
        "total": total,         # ← 正確的最終數量
        "page": page,
        "per_page": per_page,
        "query": request.query_string.decode(errors="ignore") or ""
    })


def _facet_counts_with_unknown(df: pd.DataFrame, col_key: str, top_k: int = 6) -> Dict[str, Any]:
    """Compute facet rows + unknown count, with robust handling for NaN/strings."""
    rows: List[Dict[str, Any]] = []
    unknown: int = 0

    key_to_col = {
        "ct_phase": ("__ct", str),
        "manufacturer": ("__mfr", str),
        "year": ("__year_int", int),
        "sex": ("__sex", str),
        "tumor": ("__tumor01", int),
        "model": ("model", str),
        "study_type": ("study_type", str),
        "site_nat": ("site_nationality", str),
        "site_nationality": ("site_nationality", str),
    }
    if col_key not in key_to_col:
        return {"rows": [], "unknown": 0}

    col_name, _typ = key_to_col[col_key]
    if col_name not in df.columns:
        return {"rows": [], "unknown": 0}

    ser = df[col_name]

    # ---- Year：數值化、NaN 視為 unknown ----
    if col_key == "year":
        s_num = pd.to_numeric(ser, errors="coerce")
        unknown = int(s_num.isna().sum())
        vc = s_num.dropna().astype(int).value_counts()
        rows = [{"value": int(v), "count": int(c)} for v, c in vc.items()]
        rows.sort(key=lambda x: (-x["count"], x["value"]))
        if top_k and top_k > 0:
            rows = rows[:top_k]
        return {"rows": rows, "unknown": unknown}

    # ---- 其他欄位：把空字串/unknown 類型歸入 unknown ----
    s_str = ser.astype(str).str.strip()
    s_lc = s_str.str.lower()
    unknown_mask = ser.isna() | (s_str == "") | (s_lc.isin({"unknown", "nan", "none", "n/a", "na"}))
    unknown = int(unknown_mask.sum())

    vals = ser[~unknown_mask]
    vc = vals.value_counts(dropna=False)

    tmp_rows: List[Dict[str, Any]] = []
    for v, c in vc.items():
        if col_key == "tumor":
            # tumor 僅接受 0/1
            try:
                iv = int(v)
            except Exception:
                continue
            if iv not in (0, 1):
                continue
            tmp_rows.append({"value": iv, "count": int(c)})
        else:
            tmp_rows.append({"value": v, "count": int(c)})

    # 排序：count desc，再 value 升（字串比較避免型別問題）
    tmp_rows.sort(key=lambda x: (-x["count"], str(x["value"])))
    if top_k and top_k > 0:
        tmp_rows = tmp_rows[:top_k]

    rows = tmp_rows
    return {"rows": rows, "unknown": unknown}


def _prune_zero_rows(rows: List[Dict[str, Any]], keep_zero: bool) -> List[Dict[str, Any]]:
    """依需求濾掉 count<=0；當 keep_zero=True（對應 guarantee=1）則不濾。"""
    if keep_zero:
        return rows
    out: List[Dict[str, Any]] = []
    for r in rows or []:
        try:
            c = int(r.get("count") or 0)
        except Exception:
            c = 0
        if c > 0:
            out.append(r)
    return out


@api_blueprint.route("/facets", methods=["GET"])
def api_facets():
    try:
        fields_raw = (_arg("fields","ct_phase,manufacturer") or "").strip()
        fields = [f.strip().lower() for f in fields_raw.split(",") if f.strip()]

        valid  = {
            "ct_phase","manufacturer","year","sex","tumor",
            "model","study_type","site_nat","site_nationality"
        }
        fields = [f for f in fields if f in valid] or ["ct_phase","manufacturer"]
        top_k  = to_int(_arg("top_k","6")) or 6
        guarantee = (_arg("guarantee","0") or "0").strip().lower() in ("1","true","yes","y")

        # 先應用目前的過濾條件
        df_now = apply_filters(DF)
        base_for_ranges = df_now if len(df_now) else DF

        facets: Dict[str, List[Dict[str, Any]]] = {}
        unknown_counts: Dict[str, int] = {}

        # 為每個 facet 準備自我排除的條件（避免自我影響）
        exclude_map = {
            "ct_phase": {"ct_phase"},
            "manufacturer": {"manufacturer","mfr_is_null","manufacturer_is_null"},
            "year": {"year_from","year_to"},
            "sex": {"sex"},
            "tumor": {"tumor"},
            "model": {"model"},
            "study_type": {"study_type"},
            "site_nat": {"site_nat","site_nationality"},
            "site_nationality": {"site_nat","site_nationality"},
        }

        for f in fields:
            ex = exclude_map.get(f, set())
            # 若 guarantee=1 且目前篩完為空，改用全量 DF 以「保證列出所有可能值」
            src = (DF if (guarantee and len(df_now) == 0) else df_now)
            df_facet = apply_filters(src, exclude=ex)
            res = _facet_counts_with_unknown(df_facet, f, top_k=top_k)

            # guarantee=0 時砍掉 count<=0 的項目
            rows = _prune_zero_rows(res.get("rows") or [], keep_zero=guarantee)
            facets[f] = rows
            unknown_counts[f] = int(res.get("unknown") or 0)

        # 年齡/年份範圍（原樣保留）
        def _minmax(series: pd.Series):
            s = series.dropna()
            if not len(s): return (None, None)
            return (float(s.min()), float(s.max()))

        age_min = age_max = None
        year_min = year_max = None
        if "__age" in base_for_ranges:
            age_min, age_max = _minmax(base_for_ranges["__age"])
        if "__year_int" in base_for_ranges:
            yr = base_for_ranges["__year_int"].dropna().astype(int)
            if len(yr):
                year_min, year_max = int(yr.min()), int(yr.max())

        return jsonify({
            "facets": facets,
            "unknown_counts": unknown_counts,
            "age_range": {"min": age_min, "max": age_max},
            "year_range": {"min": year_min, "max": year_max},
            "total": int(len(df_now)),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    
@api_blueprint.route("/random", methods=['GET'])
def api_random_topk_rotate_norand():
    """
    推薦：完整資料優先 → 取 Top-K(預設100) → 環狀位移 → 可排除最近看過
    排序：__spacing_sum ↑, __shape_sum ↓, __case_sortkey ↑
    """
    try:
        scope = (request.args.get("scope", "filtered") or "filtered").strip().lower()
        base_df = apply_filters(DF)
        if len(base_df) == 0 and scope == "all":
            base_df = DF.copy()

        base_df = ensure_sort_cols(base_df)

        # 只取完整資料；若沒有完整的就退回全部
        df_full = base_df[base_df["__complete"]] if "__complete" in base_df.columns else base_df
        if len(df_full) == 0:
            df_full = base_df
        df = df_full.sort_values(
            by=["__spacing_sum","__shape_sum","__case_sortkey"],
            ascending=[True, False, True],
            na_position="last",
            kind="mergesort",
        )

        if len(df) == 0:
            return jsonify({"items": [], "total": 0, "meta": {"k": 0, "used_recent": 0}}), 200

        # n, k
        try: n = int(request.args.get("n") or 3)
        except Exception: n = 3
        n = max(1, min(n, len(df)))

        try: K = int(request.args.get("k") or 100)
        except Exception: K = 100
        K = max(n, min(K, len(df)))

        # recent 排除
        recent_raw = (request.args.get("recent") or "").strip()
        used_recent = 0
        if recent_raw:
            recent_ids = {s.strip() for s in recent_raw.split(",") if s.strip()}
            key = df["__case_str"].astype(str) if "__case_str" in df.columns else None
            if key is not None:
                mask = ~key.isin(recent_ids)
                used_recent = int((~mask).sum())
                df2 = df[mask]
                if len(df2): df = df2

        topk = df.iloc[:K]
        if len(topk) == 0:
            return jsonify({"items": [], "total": 0, "meta": {"k": 0, "used_recent": used_recent}}), 200

        off_arg = request.args.get("offset")
        if off_arg is not None:
            try: offset = int(off_arg) % len(topk)
            except Exception: offset = 0
        else:
            now = datetime.utcnow()
            offset = ((now.minute * 60) + now.second) % len(topk)

        idx = list(range(len(topk))) + list(range(len(topk)))
        pick = idx[offset:offset + min(n, len(topk))]
        sub = topk.iloc[pick]

        items = [row_to_item(r) for _, r in sub.iterrows()]
        resp = jsonify({
            "items": clean_json_list(items),
            "total": int(len(df)),
            "meta": {"k": int(len(topk)), "used_recent": used_recent, "offset": int(offset)}
        })
        r = make_response(resp)
        r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        r.headers["Pragma"] = "no-cache"
        r.headers["Expires"] = "0"
        return r

    except Exception as e:
        return jsonify({"error": str(e)}), 400

AI_ALLOWED_ACTION_TYPES = {
    "isolate_organs",
    "show_organs",
    "hide_organs",
    "focus_organ",
    "get_organ_metric",
    "set_opacity",
    "set_window",
    "set_window_preset",
    "set_zoom",
    "zoom_to_fit",
    "set_view",
    "activate_measurement_tool",
    "clear_measurements",
    "list_structures",
    "get_structure_count",
    "get_largest_structure",
    "get_smallest_structure",
}
AI_ALLOWED_VIEWS = {"mpr", "axial", "sagittal", "coronal", "3d"}
AI_ALLOWED_PRESETS = {"soft_tissue", "bone", "lung", "liver"}
AI_ALLOWED_TOOLS = {"distance", "probe", "roi"}
AI_ALLOWED_METRICS = {"volume_cm3", "mean_hu", "all"}


def _ai_norm(value):
    return " ".join(str(value or "").lower().replace("_", " ").replace(".nii.gz", "").replace(".nii", "").split())


def _ai_display(value):
    text = str(value or "").replace(".nii.gz", "").replace(".nii", "").replace("_", " ").strip()
    return text.title() if text else "Structure"


def _ai_metric_valid(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return False
    return np.isfinite(number) and number > 0 and number != 999999


def _ai_public_metric(entry):
    return {
        "organ_name": str(entry.get("organ_name") or ""),
        "display_name": _ai_display(entry.get("organ_name")),
        "volume_cm3": entry.get("volume_cm3"),
        "mean_hu": entry.get("mean_hu"),
    }


def _ai_load_metrics(case_id, supplied_metrics):
    """
    Prefer server-computed segmentation metrics.

    If the server does not have local NIfTI data, use metrics supplied by
    the frontend. Never allow a missing local file to prevent Ollama from
    answering general educational questions.
    """

    identifier = str(case_id or "").strip()

    if identifier and _is_safe_id(identifier):
        try:
            if identifier.isdigit():
                result = get_mask_data_internal(identifier)
                source = "server_mask_data"
            else:
                result = calculate_session_metrics(
                    identifier,
                    Constants.SESSIONS_DIR_NAME,
                )
                source = "session_mask_data"

            if isinstance(result, dict) and not result.get("error"):
                raw_metrics = result.get("organ_metrics") or []

                if isinstance(raw_metrics, list):
                    cleaned = [
                        _ai_public_metric(item)
                        for item in raw_metrics
                        if isinstance(item, dict)
                    ]

                    if cleaned:
                        return cleaned, source

        except Exception as error:
            # Missing local case data should not disable general AI chat.
            print(
                "[AI metrics unavailable]",
                type(error).__name__,
                str(error),
            )

    if isinstance(supplied_metrics, list):
        cleaned = [
            _ai_public_metric(item)
            for item in supplied_metrics
            if isinstance(item, dict)
        ]

        if cleaned:
            return cleaned, "frontend_supplied_metrics"

    return [], "unavailable"


def _ai_metric_lookup(metrics, available_organs):
    lookup = {}
    for entry in metrics:
        organ_name = entry.get("organ_name")
        display_name = entry.get("display_name") or _ai_display(organ_name)
        for key in {organ_name, display_name, _ai_norm(organ_name), _ai_norm(display_name)}:
            if key:
                lookup[_ai_norm(key)] = entry
    for organ in available_organs:
        key = _ai_norm(organ)
        if key not in lookup:
            # Keep the viewer catalog entry available for honest “metric unavailable” responses.
            lookup[key] = {"organ_name": organ, "display_name": _ai_display(organ), "volume_cm3": None, "mean_hu": None}
    return lookup


def _ai_resolve_organ(value, available_organs):
    norm = _ai_norm(value)
    if not norm:
        return None

    normalized_available = [(organ, _ai_norm(organ)) for organ in available_organs]
    for organ, organ_norm in normalized_available:
        if norm == organ_norm:
            return organ
    for organ, organ_norm in normalized_available:
        if norm in organ_norm or organ_norm in norm:
            return organ

    alias_groups = [
        {"left kidney", "kidney left", "left renal", "left renal kidney"},
        {"right kidney", "kidney right", "right renal", "right renal kidney"},
        {"gall bladder", "gallbladder"},
        {"inferior vena cava", "ivc", "vena cava", "postcava"},
        {"superior mesenteric artery", "sma"},
        {"common bile duct", "bile duct", "cbd"},
        {"left adrenal", "left adrenal gland", "adrenal gland left"},
        {"right adrenal", "right adrenal gland", "adrenal gland right"},
    ]
    for aliases in alias_groups:
        normalized_aliases = {_ai_norm(alias) for alias in aliases}
        if norm not in normalized_aliases:
            continue
        for organ, organ_norm in normalized_available:
            if organ_norm in normalized_aliases:
                return organ
    return None


def _ai_sanitize_actions(actions, available_organs):
    if not isinstance(actions, list):
        return []
    sanitized = []
    seen = set()
    for action in actions[:8]:
        if not isinstance(action, dict):
            continue
        action_type = action.get("type")
        if action_type not in AI_ALLOWED_ACTION_TYPES:
            continue
        clean = {"type": action_type}
        if action_type in {"isolate_organs", "show_organs", "hide_organs"}:
            organs = action.get("organs")
            if not isinstance(organs, list):
                continue
            resolved = []
            for organ in organs:
                found = _ai_resolve_organ(organ, available_organs)
                if found and found not in resolved:
                    resolved.append(found)
            if not resolved:
                continue
            clean["organs"] = resolved
        elif action_type == "focus_organ":
            found = _ai_resolve_organ(action.get("organ"), available_organs)
            if not found:
                continue
            clean["organ"] = found
        elif action_type == "get_organ_metric":
            found = _ai_resolve_organ(action.get("organ"), available_organs)
            if not found:
                continue
            metric = action.get("metric") or "volume_cm3"
            if metric not in AI_ALLOWED_METRICS:
                metric = "volume_cm3"
            clean["organ"] = found
            clean["metric"] = metric
        elif action_type == "set_opacity":
            try:
                clean["value"] = max(0, min(100, float(action.get("value"))))
            except (TypeError, ValueError):
                continue
        elif action_type == "set_window":
            try:
                clean["width"] = max(1, float(action.get("width")))
                clean["center"] = float(action.get("center"))
            except (TypeError, ValueError):
                continue
        elif action_type == "set_window_preset":
            preset = action.get("preset")
            if preset not in AI_ALLOWED_PRESETS:
                continue
            clean["preset"] = preset
        elif action_type == "set_zoom":
            try:
                clean["value"] = max(0.1, min(20, float(action.get("value"))))
            except (TypeError, ValueError):
                continue
        elif action_type == "set_view":
            view = action.get("view")
            if view not in AI_ALLOWED_VIEWS:
                continue
            clean["view"] = view
        elif action_type == "activate_measurement_tool":
            tool = action.get("tool")
            if tool not in AI_ALLOWED_TOOLS:
                continue
            clean["tool"] = tool

        key = json.dumps(clean, sort_keys=True)
        if key not in seen:
            sanitized.append(clean)
            seen.add(key)
    return sanitized


def _ai_action_family(action_type):
    if action_type in {"isolate_organs", "show_organs", "hide_organs", "focus_organ"}:
        return "organ_visibility"
    if action_type in {"set_window", "set_window_preset"}:
        return "window"
    if action_type in {"set_zoom", "zoom_to_fit"}:
        return "zoom"
    if action_type in {"activate_measurement_tool", "clear_measurements"}:
        return "measurement_tool"
    if action_type in {"list_structures", "get_structure_count", "get_largest_structure", "get_smallest_structure"}:
        return "structure_query"
    return action_type


def _ai_merge_actions(deterministic_actions, model_actions):
    """Prefer deterministic actions for recognized commands and let Ollama fill gaps."""
    merged = []
    seen = set()
    used_families = set()
    for action in [*(deterministic_actions or []), *(model_actions or [])]:
        if not isinstance(action, dict):
            continue
        key = json.dumps(action, sort_keys=True)
        if key in seen:
            continue
        family = _ai_action_family(action.get("type"))
        # A command should not apply conflicting visibility/window/zoom/tool actions.
        if family in used_families and family != "get_organ_metric":
            continue
        merged.append(action)
        seen.add(key)
        used_families.add(family)
    return merged[:8]


def _ai_metadata(case_id, supplied):
    metadata = {}
    if isinstance(supplied, dict):
        for key in ["sex", "age", "bmi", "height_cm", "weight_kg"]:
            if supplied.get(key) not in [None, ""]:
                metadata[key] = supplied.get(key)
    if case_id and str(case_id).isdigit():
        entry = _METADATA_CACHE.get(get_panTS_id(str(case_id)), {})
        if entry.get("age") not in [None, ""]:
            try:
                metadata["age"] = float(entry.get("age"))
            except (TypeError, ValueError):
                metadata["age"] = entry.get("age")
        if entry.get("sex") not in [None, ""]:
            metadata["sex"] = entry.get("sex")
    if "bmi" not in metadata:
        try:
            height_m = float(metadata.get("height_cm")) / 100
            weight_kg = float(metadata.get("weight_kg"))
            if height_m > 0 and weight_kg > 0:
                metadata["bmi"] = round(weight_kg / (height_m * height_m), 1)
        except (TypeError, ValueError):
            pass
    return metadata


def _ai_has_case_reference(norm: str) -> bool:
    case_phrases = (
        "this scan",
        "this ct",
        "this case",
        "this patient",
        "this segmentation",
        "the segmentation",
        "current scan",
        "current case",
        "currently loaded",
        "shown here",
        "in the viewer",
        "in this image",
        "in these images",
        "my scan",
        "my ct",
        "my liver",
        "my pancreas",
        "my kidney",
        "my spleen",
        "do i have",
        "am i",
        "for this patient",
        "patient s",
        "patient's",
    )

    return any(phrase in norm for phrase in case_phrases)


def _ai_question_mode(message, fallback_actions):
    """
    Separate general knowledge from case-specific questions.

    Modes:
    - general_education
    - case_metadata
    - case_measurement
    - case_health_context
    - viewer_command
    """

    norm = _ai_norm(message)
    action_types = {
        action.get("type")
        for action in (fallback_actions or [])
        if isinstance(action, dict)
    }

    viewer_action_types = {
        "isolate_organs",
        "show_organs",
        "hide_organs",
        "focus_organ",
        "set_opacity",
        "set_window",
        "set_window_preset",
        "set_zoom",
        "zoom_to_fit",
        "set_view",
        "activate_measurement_tool",
        "clear_measurements",
    }

    if action_types.intersection(viewer_action_types):
        return "viewer_command"

    has_case_reference = _ai_has_case_reference(norm)

    if "bmi" in norm or "body mass index" in norm:
        case_bmi_phrases = (
            "patient bmi",
            "patient s bmi",
            "patient's bmi",
            "this bmi",
            "their bmi",
            "his bmi",
            "her bmi",
            "my bmi",
            "bmi for this",
            "bmi of this",
        )

        if has_case_reference or any(
            phrase in norm for phrase in case_bmi_phrases
        ):
            return "case_metadata"

        return "general_education"

    if "age" in norm or "how old" in norm:
        if has_case_reference or "patient age" in norm:
            return "case_metadata"

    health_terms = (
        "healthy",
        "unhealthy",
        "normal",
        "abnormal",
        "concerning",
        "disease",
        "diseased",
        "cancer",
        "cancerous",
        "tumor",
        "malignant",
        "benign",
        "enlarged",
        "too large",
        "too small",
        "swollen",
        "damaged",
        "cirrhosis",
        "fatty liver",
        "lesion",
        "mass",
    )

    health_question = any(
        term in norm for term in health_terms
    )

    implied_current_case = (
        norm.startswith("is the liver ")
        or norm.startswith("is the pancreas ")
        or norm.startswith("is the spleen ")
        or norm.startswith("is the kidney ")
        or norm.startswith("does the liver look ")
    )

    if health_question and (
        has_case_reference or implied_current_case
    ):
        return "case_health_context"

    measurement_terms = (
        "volume",
        "how big",
        "what size",
        "size of",
        "mean hu",
        "hounsfield",
        "largest structure",
        "smallest structure",
        "how many structures",
        "structure count",
        "measured",
    )

    asks_measurement = any(
        term in norm for term in measurement_terms
    )

    general_reference_terms = (
        "normal liver volume",
        "normal organ volume",
        "typical liver volume",
        "average liver volume",
        "usual liver size",
        "what is considered normal",
    )

    if asks_measurement:
        if (
            any(term in norm for term in general_reference_terms)
            and not has_case_reference
        ):
            return "general_education"

        return "case_measurement"

    return "general_education"


def _ai_case_metadata_reply(norm, metadata):
    """
    Answer only explicitly case-specific age/BMI questions.

    General questions such as 'What is BMI?' must continue to Ollama.
    """

    has_case_reference = _ai_has_case_reference(norm)

    asks_case_age = (
        ("age" in norm or "how old" in norm)
        and (
            has_case_reference
            or "patient age" in norm
            or "age of the patient" in norm
        )
    )

    if asks_case_age:
        if metadata.get("age") not in [None, ""]:
            age = metadata.get("age")

            try:
                age = round(float(age))
            except (TypeError, ValueError):
                pass

            return (
                "The available metadata lists this patient's age "
                f"as **{age} years**."
            )

        return (
            "The current case metadata does not include a valid age."
        )

    asks_case_bmi = (
        ("bmi" in norm or "body mass index" in norm)
        and (
            has_case_reference
            or "patient bmi" in norm
            or "patient's bmi" in norm
            or "bmi of the patient" in norm
            or "bmi for the patient" in norm
        )
    )

    if asks_case_bmi:
        if metadata.get("bmi") not in [None, ""]:
            try:
                bmi = float(metadata.get("bmi"))
                return (
                    "The available metadata lists this patient's BMI "
                    f"as **{bmi:.1f}**."
                )
            except (TypeError, ValueError):
                pass

        return (
            "The current case does not include enough height and weight "
            "information to calculate the patient's BMI."
        )

    return None


def _ai_relevant_metrics(message, metrics):
    norm = _ai_norm(message)
    relevant = []

    for entry in metrics or []:
        if not isinstance(entry, dict):
            continue

        organ_name = entry.get("organ_name")
        display_name = (
            entry.get("display_name")
            or _ai_display(organ_name)
        )

        names = {
            _ai_norm(organ_name),
            _ai_norm(display_name),
        }

        if any(name and name in norm for name in names):
            relevant.append(entry)

    return relevant


def _ai_case_facts(message, metrics, metadata):
    """
    Produce exact deterministic facts that can be placed beside the model's
    health-context explanation.
    """

    facts = []
    relevant_metrics = _ai_relevant_metrics(
        message,
        metrics,
    )

    for entry in relevant_metrics:
        organ = (
            entry.get("display_name")
            or _ai_display(entry.get("organ_name"))
        )

        volume = entry.get("volume_cm3")
        mean_hu = entry.get("mean_hu")

        organ_facts = []

        if _ai_metric_valid(volume):
            organ_facts.append(
                f"segmented volume {float(volume):.2f} cm³"
            )

        if _ai_metric_valid(mean_hu) or mean_hu == 0:
            organ_facts.append(
                f"mean attenuation {float(mean_hu):.1f} HU"
            )

        if organ_facts:
            facts.append(
                f"**{organ}:** " + ", ".join(organ_facts)
            )

    if metadata.get("age") not in [None, ""]:
        facts.append(f"**Age:** {metadata.get('age')}")

    if metadata.get("sex") not in [None, ""]:
        facts.append(f"**Sex:** {metadata.get('sex')}")

    if metadata.get("bmi") not in [None, ""]:
        facts.append(f"**BMI:** {metadata.get('bmi')}")

    return facts


def _ai_extract_unknown_volume_request(norm, metric_lookup):
    if "volume" not in norm and "how big" not in norm and "size" not in norm:
        return None
    for key, entry in metric_lookup.items():
        if key and key in norm:
            return entry
    # A small deny-list for common out-of-scope organs people may ask about.
    for organ in ["brain", "heart", "eye", "skull"]:
        if organ in norm:
            return {"organ_name": organ, "display_name": _ai_display(organ), "volume_cm3": None, "mean_hu": None, "missing": True}
    return None


def _ai_action_prefix(actions):
    bits = []
    if any(a.get("type") == "set_view" and a.get("view") == "3d" for a in actions):
        bits.append("switched to 3D")
    isolate = next((a for a in actions if a.get("type") == "isolate_organs"), None)
    if isolate:
        bits.append("isolated " + ", ".join(_ai_display(o) for o in isolate.get("organs", [])))
    elif any(a.get("type") == "show_organs" for a in actions):
        show = next(a for a in actions if a.get("type") == "show_organs")
        bits.append("showed " + ", ".join(_ai_display(o) for o in show.get("organs", [])))
    if any(a.get("type") == "focus_organ" for a in actions):
        focus = next(a for a in actions if a.get("type") == "focus_organ")
        bits.append("focused on " + _ai_display(focus.get("organ")))
    if bits:
        if len(bits) == 1:
            return "I " + bits[0] + ". "
        return "I " + ", ".join(bits[:-1]) + ", and " + bits[-1] + ". "
    return ""


def _ai_structure_names(metrics, available_organs):
    raw_names = [m.get("organ_name") for m in metrics if m.get("organ_name")]
    if not raw_names:
        raw_names = available_organs
    names = []
    seen = set()
    for name in raw_names:
        display = _ai_display(name)
        key = _ai_norm(display)
        if key and key not in seen:
            names.append(display)
            seen.add(key)
    return names


def _ai_action_confirmation(actions):
    confirmations = []
    for action in actions:
        action_type = action.get("type")
        if action_type == "isolate_organs":
            confirmations.append("Isolated " + ", ".join(_ai_display(o) for o in action.get("organs", [])) + ".")
        elif action_type == "show_organs":
            confirmations.append("Showed " + ", ".join(_ai_display(o) for o in action.get("organs", [])) + ".")
        elif action_type == "hide_organs":
            confirmations.append("Hid " + ", ".join(_ai_display(o) for o in action.get("organs", [])) + ".")
        elif action_type == "focus_organ":
            confirmations.append("Focused on " + _ai_display(action.get("organ")) + ".")
        elif action_type == "set_view":
            confirmations.append(f"Switched to {str(action.get('view')).upper()} view.")
        elif action_type == "set_opacity":
            confirmations.append(f"Set overlay opacity to {float(action.get('value')):.0f}%.")
        elif action_type == "set_window_preset":
            confirmations.append(f"Applied the {str(action.get('preset')).replace('_', ' ')} window preset.")
        elif action_type == "set_window":
            confirmations.append(f"Set the CT window to width {float(action.get('width')):.0f} and center {float(action.get('center')):.0f}.")
        elif action_type == "set_zoom":
            confirmations.append(f"Set zoom to {float(action.get('value')):g}.")
        elif action_type == "zoom_to_fit":
            confirmations.append("Reset zoom to fit.")
        elif action_type == "activate_measurement_tool":
            confirmations.append(f"Activated the {str(action.get('tool')).upper()} tool.")
        elif action_type == "clear_measurements":
            confirmations.append("Cleared the current measurements.")
    return " ".join(confirmations)


def _ai_grounded_reply(
    message,
    actions,
    metrics,
    available_organs,
    metadata,
    candidate_reply,
    question_mode,
):
    norm = _ai_norm(message)
    metric_lookup = _ai_metric_lookup(
        metrics,
        available_organs,
    )

    # Only intercept age/BMI when the user explicitly asks about
    # the currently loaded patient.
    if question_mode == "case_metadata":
        metadata_reply = _ai_case_metadata_reply(
            norm,
            metadata,
        )

        if metadata_reply:
            return metadata_reply

    # For health questions, preserve Ollama's explanation while placing
    # exact measured case facts above it.
    if question_mode == "case_health_context":
        facts = _ai_case_facts(
            message,
            metrics,
            metadata,
        )

        if (
            isinstance(candidate_reply, str)
            and candidate_reply.strip()
        ):
            explanation = candidate_reply.strip()
        else:
            explanation = (
                "The available segmentation measurements can provide "
                "useful context, but they are not enough by themselves "
                "to determine whether this organ is healthy. A complete "
                "assessment would also consider the CT appearance, "
                "attenuation, enhancement, contour, focal lesions, "
                "clinical history, and relevant laboratory results."
            )

        if facts:
            reply = (
                "Available measurements for this case:\n"
                + "\n".join(f"- {fact}" for fact in facts)
                + "\n\n"
                + explanation
            )
        else:
            reply = explanation

        uncertainty_terms = (
            "not a diagnosis",
            "cannot confirm",
            "cannot determine",
            "volume alone",
            "measurements alone",
            "radiologist",
            "clinical evaluation",
        )

        if not any(
            term in explanation.lower()
            for term in uncertainty_terms
        ):
            reply += (
                "\n\nThis is an educational, non-diagnostic assessment. "
                "Volume and segmentation measurements alone cannot "
                "confirm that an organ is healthy or diagnose disease."
            )

        confirmation = _ai_action_confirmation(actions)

        if confirmation:
            reply += "\n\n" + confirmation

        return reply

    # Deterministically ground exact case measurements.
    for action in actions:
        if action.get("type") != "get_organ_metric":
            continue

        entry = metric_lookup.get(
            _ai_norm(action.get("organ"))
        )

        organ_label = _ai_display(
            action.get("organ")
        )

        if not entry or entry.get("missing"):
            return (
                f"{organ_label} is not available in this segmentation, "
                f"so no {organ_label.lower()} measurement is available "
                "for this case."
            )

        metric = action.get("metric") or "volume_cm3"
        volume = entry.get("volume_cm3")
        mean_hu = entry.get("mean_hu")
        prefix = _ai_action_prefix(actions)

        if metric == "mean_hu":
            if _ai_metric_valid(mean_hu) or mean_hu == 0:
                return (
                    f"{prefix}The segmented **{organ_label}** mean "
                    f"attenuation is **{float(mean_hu):.1f} HU**, "
                    "measured from the segmentation mask."
                )

            return (
                f"{prefix}Mean HU for **{organ_label}** is not "
                "available for this case."
            )

        if metric == "all":
            parts = []

            if _ai_metric_valid(volume):
                parts.append(
                    f"volume **{float(volume):.2f} cm³**"
                )

            if _ai_metric_valid(mean_hu) or mean_hu == 0:
                parts.append(
                    f"mean attenuation "
                    f"**{float(mean_hu):.1f} HU**"
                )

            if parts:
                return (
                    f"{prefix}For the segmented **{organ_label}**, "
                    + " and ".join(parts)
                    + "."
                )

            return (
                f"{prefix}Metrics for **{organ_label}** are not "
                "available for this case."
            )

        if _ai_metric_valid(volume):
            return (
                f"{prefix}The segmented **{organ_label}** volume is "
                f"**{float(volume):.2f} cm³**, measured from the "
                "segmentation mask."
            )

        return (
            f"{prefix}No valid segmented volume is available for "
            f"**{organ_label}** in this case."
        )

    if any(
        action.get("type") == "get_largest_structure"
        for action in actions
    ):
        valid = [
            metric
            for metric in metrics
            if _ai_metric_valid(metric.get("volume_cm3"))
        ]

        if valid:
            largest = max(
                valid,
                key=lambda item: float(
                    item.get("volume_cm3")
                ),
            )

            return (
                "The largest segmented structure is "
                f"**{_ai_display(largest.get('organ_name'))}**, "
                "with a measured volume of "
                f"**{float(largest.get('volume_cm3')):.2f} cm³**."
            )

        return (
            "I could not determine the largest structure because valid "
            "volume metrics are unavailable for this case."
        )

    if any(
        action.get("type") == "get_smallest_structure"
        for action in actions
    ):
        valid = [
            metric
            for metric in metrics
            if _ai_metric_valid(metric.get("volume_cm3"))
        ]

        if valid:
            smallest = min(
                valid,
                key=lambda item: float(
                    item.get("volume_cm3")
                ),
            )

            return (
                "The smallest segmented structure is "
                f"**{_ai_display(smallest.get('organ_name'))}**, "
                "with a measured volume of "
                f"**{float(smallest.get('volume_cm3')):.2f} cm³**."
            )

        return (
            "I could not determine the smallest structure because valid "
            "volume metrics are unavailable for this case."
        )

    if any(
        action.get("type") == "get_structure_count"
        for action in actions
    ):
        names = _ai_structure_names(
            metrics,
            available_organs,
        )

        return (
            f"This case has **{len(names)} segmented structures** "
            "listed for the viewer."
        )

    if any(
        action.get("type") == "list_structures"
        for action in actions
    ):
        names = _ai_structure_names(
            metrics,
            available_organs,
        )

        if names:
            return (
                f"This case includes **{len(names)} segmented "
                "structures**: "
                + ", ".join(names)
                + "."
            )

        return (
            "No segmented structures are listed for this case."
        )

    unknown_volume = _ai_extract_unknown_volume_request(
        norm,
        metric_lookup,
    )

    if unknown_volume and unknown_volume.get("missing"):
        organ_name = _ai_display(
            unknown_volume.get("organ_name")
        )

        return (
            f"The **{organ_name}** is not included in this abdominal "
            f"segmentation, so no {organ_name.lower()} volume is "
            "available for this case."
        )

    confirmation = _ai_action_confirmation(actions)

    if (
        isinstance(candidate_reply, str)
        and candidate_reply.strip()
    ):
        reply = candidate_reply.strip()

        if confirmation:
            reply += "\n\n" + confirmation

        return reply

    if confirmation:
        return confirmation

    return (
        "I could not generate a complete response. Please verify that "
        "Ollama is running and that the selected model is installed."
    )


def _ai_system_prompt():
    return """
You are BodyMaps AI, a capable and conversational assistant embedded in
an abdominal CT visualization application.

Your job is to answer the user's actual question. Do not assume that
every question asks about the current patient.

You operate in several modes.

GENERAL EDUCATIONAL QUESTIONS
Answer ordinary questions using your general knowledge.

Examples:
- What is a CT scan?
- What is body mass index?
- What does the liver do?
- What are Hounsfield units?
- What is a segmentation mask?
- What conditions can affect the liver?
- What is a typical liver volume?

These questions do not require patient metadata. Answer them naturally,
clearly, and conversationally.

CASE-SPECIFIC MEASUREMENTS
When the user asks about this scan, this case, this patient, the current
segmentation, or a measured structure, use the supplied case data.

Examples:
- What is the liver volume in this scan?
- How big is the segmented pancreas?
- What is this patient's BMI?
- Which segmented structure is largest?

Never invent a case-specific value. Use exact supplied values.

CASE-SPECIFIC HEALTH QUESTIONS
When the user asks questions such as:
- Is this liver healthy?
- Does this liver appear enlarged?
- Is this organ normal?
- Is this finding concerning?
- Could this be a tumor?

Provide a useful evidence-limited assessment rather than a generic
refusal.

You should:
1. State the exact available case measurements.
2. Explain what those measurements may suggest.
3. Compare them with broad educational expectations when appropriate,
   while clearly labeling those comparisons as approximate.
4. Explain important limitations of the available evidence.
5. State what additional imaging features, metadata, laboratory values,
   symptoms, or clinical history would normally be considered.

Do not state that a patient definitely has or does not have a disease
when the supplied data does not establish that conclusion.

Volume alone does not prove that an organ is healthy or unhealthy.
A segmented volume can provide useful context, but complete assessment
may also require contour, attenuation, contrast enhancement, focal
lesions, surrounding structures, prior scans, laboratory results, and
clinical history.

You may explain possible diseases and general treatment approaches.
Do not prescribe medication, choose a personalized treatment, or tell
the user to ignore professional medical care.

VIEWER COMMANDS
Translate viewer requests into the allowed structured actions.
Actions are applied immediately and should not require confirmation.

GROUNDING RULES
- Never invent organ volume, mean HU, age, BMI, or patient metadata.
- General medical and imaging knowledge does not need to be present in
  the case metadata.
- Patient-specific facts must come from supplied data.
- If case measurements are unavailable, say which data is missing and
  still answer the educational portion of the question.
- "Segment the liver" means display or isolate an existing segmentation.
  Do not claim a new segmentation model was run unless the backend
  actually reports that it ran.

OUTPUT FORMAT
Return exactly one JSON object:

{
  "reply": "complete conversational answer",
  "actions": [],
  "intent": "short_intent_name"
}

Allowed actions:
- {"type":"isolate_organs","organs":["exact available organ name"]}
- {"type":"show_organs","organs":["exact available organ name"]}
- {"type":"hide_organs","organs":["exact available organ name"]}
- {"type":"focus_organ","organ":"exact available organ name"}
- {"type":"get_organ_metric","organ":"exact available organ name","metric":"volume_cm3|mean_hu|all"}
- {"type":"list_structures"}
- {"type":"get_structure_count"}
- {"type":"get_largest_structure"}
- {"type":"get_smallest_structure"}
- {"type":"set_view","view":"mpr|axial|sagittal|coronal|3d"}
- {"type":"set_opacity","value":0-100}
- {"type":"set_window","width":number,"center":number}
- {"type":"set_window_preset","preset":"soft_tissue|bone|lung|liver"}
- {"type":"set_zoom","value":number}
- {"type":"zoom_to_fit"}
- {"type":"activate_measurement_tool","tool":"distance|probe|roi"}
- {"type":"clear_measurements"}

For a question that does not need a viewer action, return an empty
actions list.

Return JSON only.
""".strip()


@api_blueprint.route("/ai-models", methods=["GET"])
def ai_models():
    try:
        models = list_ollama_models()
        model_names = [model["name"] for model in models]
        default_model = DEFAULT_OLLAMA_MODEL if DEFAULT_OLLAMA_MODEL in model_names else (model_names[0] if model_names else DEFAULT_OLLAMA_MODEL)
        return jsonify({
            "available": True,
            "models": models,
            "default_model": default_model,
        })
    except OllamaUnavailable as error:
        return jsonify({
            "available": False,
            "models": [],
            "default_model": DEFAULT_OLLAMA_MODEL,
            "error": f"Ollama is not reachable at the configured local endpoint: {error}",
        }), 200


@api_blueprint.route("/ai-command", methods=["POST"])
def ai_command():
    try:
        body = request.get_json(
            force=True,
            silent=True,
        ) or {}

        message = str(
            body.get("message") or ""
        ).strip()

        if not message:
            return jsonify(
                {
                    "reply": (
                        "Please type a question or viewer command."
                    ),
                    "actions": [],
                    "source": "validation",
                }
            ), 400

        available_organs = body.get(
            "available_organs"
        ) or []

        if not isinstance(available_organs, list):
            available_organs = []

        available_organs = [
            str(item).strip()
            for item in available_organs
            if str(item).strip()
        ]

        viewer_state = (
            body.get("viewer_state")
            if isinstance(
                body.get("viewer_state"),
                dict,
            )
            else {}
        )

        case_id = str(
            body.get("session_id")
            or body.get("case_id")
            or ""
        ).strip()

        requested_model = body.get("model")

        # Always use the configured default when the frontend does not
        # explicitly send a model.
        selected_model = (
            requested_model.strip()
            if isinstance(requested_model, str)
            and requested_model.strip()
            else DEFAULT_OLLAMA_MODEL
        )

        metrics, metric_source = _ai_load_metrics(
            case_id,
            body.get("organ_metrics"),
        )

        metadata = _ai_metadata(
            case_id,
            body.get("demographics"),
        )

        # Include metric organ names even if the frontend organ catalog
        # was empty or incomplete.
        for metric in metrics:
            organ_name = str(
                metric.get("organ_name") or ""
            ).strip()

            if (
                organ_name
                and organ_name not in available_organs
            ):
                available_organs.append(organ_name)

        fallback = parse_intent(
            message=message,
            available_organs=available_organs,
            viewer_state=viewer_state,
            case_id=case_id or None,
        )

        fallback_actions = _ai_sanitize_actions(
            fallback.get("actions", []),
            available_organs,
        )

        question_mode = _ai_question_mode(
            message,
            fallback_actions,
        )

        # Do not use the fallback's medical refusal as the model's answer.
        # The model receives the fallback only as an action suggestion.
        rule_suggestion_reply = None

        if question_mode in {
            "viewer_command",
            "case_measurement",
        }:
            rule_suggestion_reply = fallback.get("reply")

        prompt_payload = {
            "request_mode": question_mode,
            "user_message": message,
            "current_case": {
                "case_id": case_id or None,
                "available_organs": available_organs,
                "computed_organ_metrics": metrics,
                "metadata": metadata,
                "metrics_source": metric_source,
            },
            "viewer_state": viewer_state,
            "rule_based_suggestion": {
                "reply": rule_suggestion_reply,
                "actions": fallback_actions,
                "intent": fallback.get("intent"),
            },
            "response_requirements": {
                "answer_general_questions_using_model_knowledge": True,
                "use_exact_case_values_when_case_specific": True,
                "do_not_invent_patient_specific_values": True,
                "provide_non_diagnostic_health_context": True,
                "return_json_only": True,
            },
        }

        model_result = None
        model_error = None
        source = "ollama"

        try:
            model_result = chat_json(
                model=selected_model,
                system_prompt=_ai_system_prompt(),
                user_prompt=json.dumps(
                    prompt_payload,
                    ensure_ascii=False,
                ),
                temperature=0.2,
            )
        except (
            OllamaUnavailable,
            Exception,
        ) as error:
            # Keep viewer actions usable if Ollama is temporarily offline.
            model_error = str(error)
            model_result = None
            source = "rule_fallback"

            print(
                "[Ollama unavailable]",
                type(error).__name__,
                model_error,
            )

        if isinstance(model_result, dict):
            model_actions = _ai_sanitize_actions(
                model_result.get("actions", []),
                available_organs,
            )

            actions = _ai_merge_actions(
                fallback_actions,
                model_actions,
            )

            candidate_reply = (
                model_result.get("reply")
                or rule_suggestion_reply
            )

            intent = (
                model_result.get("intent")
                or fallback.get("intent")
                or question_mode
            )
        else:
            actions = fallback_actions
            candidate_reply = rule_suggestion_reply
            intent = (
                fallback.get("intent")
                or question_mode
            )

        reply = _ai_grounded_reply(
            message=message,
            actions=actions,
            metrics=metrics,
            available_organs=available_organs,
            metadata=metadata,
            candidate_reply=candidate_reply,
            question_mode=question_mode,
        )

        response = {
            "reply": reply,
            "actions": actions,
            "grounding": {
                "case_id": case_id,
                "request_mode": question_mode,
                "metrics_source": metric_source,
                "organ_count": (
                    len(metrics)
                    if metrics
                    else len(available_organs)
                ),
                "metadata_fields": sorted(
                    metadata.keys()
                ),
            },
            "source": source,
            "model": (
                selected_model
                if source == "ollama"
                else None
            ),
            "intent": intent,
        }

        if model_error:
            response["ollama_error"] = model_error

        return jsonify(response)

    except Exception as error:
        print(
            "[ai_command error]",
            type(error).__name__,
            str(error),
        )

        return jsonify(
            {
                "reply": (
                    "An internal error occurred while processing "
                    "the AI request."
                ),
                "actions": [],
                "source": "error",
                "error_type": type(error).__name__,
            }
        ), 500

# ---------------------------------------------------------------------------
# Edited segmentation masks (viewer's Edit Masks panel).
# Strictly additive and isolated: writes ONLY into {PANTS_PATH}/edited_masks/,
# never touching image_only/ or mask_only/, so the original dataset is safe.
# Each save is timestamped rather than overwritten — a lightweight version
# history a maintainer can inspect or promote manually.
# ---------------------------------------------------------------------------

_EDITED_MASKS_DIRNAME = "edited_masks"
_EDITED_MASK_MAX_BYTES = 512 * 1024 * 1024  # generous cap for a full-body labelmap


def _edited_masks_dir(case_id):
    # Traversal safety: require digits, then convert to int before it reaches the
    # filesystem. A number can't carry a "../" or "/" payload, so the only value
    # that flows into os.path.join is fully controlled (get_panTS_id just zero-pads
    # and prefixes "PanTS_"). The int() cast is also what lets static analysis
    # (CodeQL py/path-injection) see the user-tainted string is neutralized.
    if not str(case_id).isdigit():
        raise ValueError("case_id must be numeric")
    return os.path.join(Constants.PANTS_PATH, _EDITED_MASKS_DIRNAME, get_panTS_id(int(case_id)))


@api_blueprint.route('/save-edited-mask/<case_id>', methods=['POST'])
def save_edited_mask(case_id):
    try:
        uploaded = request.files.get("mask")
        if uploaded is None:
            return jsonify({"error": "No file field 'mask' in the request."}), 400
        data = uploaded.read(_EDITED_MASK_MAX_BYTES + 1)
        if len(data) > _EDITED_MASK_MAX_BYTES:
            return jsonify({"error": "Mask file too large."}), 413
        # The viewer always sends gzipped NIfTI; check the gzip magic bytes.
        if len(data) < 2 or data[0] != 0x1F or data[1] != 0x8B:
            return jsonify({"error": "Expected a gzipped NIfTI (.nii.gz) file."}), 400
        
        out_dir = _edited_masks_dir(case_id)
        os.makedirs(out_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"combined_labels_edited_{timestamp}.nii.gz"
        with open(os.path.join(out_dir, filename), "wb") as f:
            f.write(data)
        # Side car carrying custom class names/colors
        # the .nii.gz only stores integer labels, so this saves user-created class names/colors
        labels_field = request.files.get("labels")
        if labels_field is not None:
            labels_data = labels_field.read(1024 * 1024)
            try:
                parsed = json.loads(labels_data.decode("utf-8"))
            except Exception:
                parsed = None
            if isinstance(parsed, dict):
                labels_filename = filename.replace(".nii.gz", "_labels.json")
                with open(os.path.join(out_dir, labels_filename), "w") as f:
                    json.dump(parsed, f)

        return jsonify({"saved": True, "filename": filename, "bytes": len(data)})
    except Exception as error:
        print("[save_edited_mask error]", type(error).__name__, error)
        return jsonify({"error": "Failed to save the edited mask."}), 500


@api_blueprint.route('/list-edited-masks/<case_id>', methods=['GET'])
def list_edited_masks(case_id):
    try:
        out_dir = _edited_masks_dir(case_id)
        if not os.path.isdir(out_dir):
            return jsonify({"items": []})
        items = []
        for name in sorted(os.listdir(out_dir), reverse=True):
            path = os.path.join(out_dir, name)
            if not os.path.isfile(path):
                continue
            items.append({
                "filename": name,
                "bytes": os.path.getsize(path),
                "modified": datetime.fromtimestamp(os.path.getmtime(path)).isoformat(),
            })
        return jsonify({"items": items})
    except Exception as error:
        print("[list_edited_masks error]", type(error).__name__, error)
        return jsonify({"error": "Failed to list edited masks."}), 500


# ---------------------------------------------------------------------------
# Advanced analysis (viewer's AI-segment tool + vessel CPR panel).
# Additive and read-only against the dataset — loads image_only/ and mask_only/
# but writes nothing there. Heavy numeric work lives in services/advanced_analysis.
# ---------------------------------------------------------------------------

import threading

# Each of these requests loads a CT volume into worker RAM, so unbounded
# concurrency is an OOM waiting to happen. Cap in-flight analyses per process
# (gunicorn workers each get their own slots); extras get an immediate 503
# instead of queueing until the worker starves.
_ANALYSIS_SLOTS = threading.BoundedSemaphore(2)
_ANALYSIS_BUSY_RESPONSE = (
    {"error": "The analysis service is busy — try again in a moment."},
    503,
)

def _safe_case_id(case_id):
    # Traversal safety: require digits, then hand get_panTS_id an int so the
    # user value can't carry a "../" or "/" payload into the CT/mask path. The
    # int() cast is also the barrier CodeQL recognizes as sanitizing the taint.
    if not str(case_id).isdigit():
        raise ValueError("case_id must be numeric")
    return int(case_id)


def _case_ct_path(case_id, low=False):
    case_dir = f"{Constants.PANTS_PATH}/image_only/{get_panTS_id(_safe_case_id(case_id))}"
    path = f"{case_dir}/{Constants.MAIN_NIFTI_FILENAME}"
    if low:
        low_path = path.replace('.nii.gz', '_lowres.nii.gz')
        if os.path.exists(low_path):
            return low_path
    return path


def _case_mask_path(case_id, low=False):
    case_dir = f"{Constants.PANTS_PATH}/mask_only/{get_panTS_id(_safe_case_id(case_id))}"
    path = f"{case_dir}/{Constants.COMBINED_LABELS_NIFTI_FILENAME}"
    if low:
        low_path = path.replace('.nii.gz', '_lowres.nii.gz')
        if os.path.exists(low_path):
            return low_path
    return path


@api_blueprint.route('/interactive-segment/<case_id>', methods=['POST'])
def interactive_segment(case_id):
    """Click-to-segment: seed prompt -> proposed mask (.nii.gz in CT geometry).

    Body JSON: { point_lps:[x,y,z] | point_ijk:[i,j,k], tolerance?, box_lps?,
                 res?: "low"|"full" }. res should match the resolution the viewer
                 loaded so the returned mask's voxel grid aligns with the labelmap.
    """
    if not _ANALYSIS_SLOTS.acquire(blocking=False):
        return jsonify(_ANALYSIS_BUSY_RESPONSE[0]), _ANALYSIS_BUSY_RESPONSE[1]
    try:
        import numpy as np
        from services.advanced_analysis import segment_from_prompt
        body = request.get_json(force=True, silent=True) or {}
        low = (body.get("res") or "low").lower() == "low"
        ct_path = _case_ct_path(case_id, low=low)
        if not os.path.exists(ct_path):
            return jsonify({"error": "CT not found for this case on the server."}), 404

        ct_obj = nib.load(ct_path)
        # float32: half the RAM of nibabel's float64 default — these are public
        # endpoints and a full-res CT at float64 is multiple GB per request.
        ct = ct_obj.get_fdata(dtype=np.float32)
        mask = segment_from_prompt(ct, ct_obj.affine, body)
        if int(mask.sum()) == 0:
            return jsonify({"error": "Nothing grew from that point — try a different spot or a higher tolerance."}), 422

        out = nib.Nifti1Image(mask, ct_obj.affine, ct_obj.header)
        out.header.set_data_dtype('uint8')
        # nibabel serializes an uncompressed .nii to bytes; gzip it ourselves.
        import gzip as _gzip
        gz = _gzip.compress(out.to_bytes())
        resp = make_response(gz)
        resp.headers['Content-Type'] = 'application/gzip'
        resp.headers['X-Mask-Voxels'] = str(int(mask.sum()))
        resp.headers['Cross-Origin-Resource-Policy'] = 'cross-origin'
        return resp
    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400
    except Exception as error:
        print("[interactive_segment error]", type(error).__name__, error)
        return jsonify({"error": "Interactive segmentation failed."}), 500
    finally:
        _ANALYSIS_SLOTS.release()


@api_blueprint.route('/vessel-cpr/<case_id>', methods=['POST'])
def vessel_cpr(case_id):
    """Straightened vessel reformat + tumour-contact metrics for staging.

    Body JSON: { vessel_label:int, lesion_label?:int (default 22 pancreatic
                 lesion), res?, slab_radius_mm?, window? }. Returns metrics +
                 the reformat as a base64 PNG (already windowed for display).
    """
    if not _ANALYSIS_SLOTS.acquire(blocking=False):
        return jsonify(_ANALYSIS_BUSY_RESPONSE[0]), _ANALYSIS_BUSY_RESPONSE[1]
    try:
        import numpy as np
        from services.advanced_analysis import analyze_vessel
        body = request.get_json(force=True, silent=True) or {}
        vessel_label = int(body.get("vessel_label", 0))
        if vessel_label <= 0:
            return jsonify({"error": "vessel_label is required."}), 400
        lesion_label = int(body.get("lesion_label", 22))
        low = (body.get("res") or "low").lower() == "low"

        ct_path = _case_ct_path(case_id, low=low)
        mask_path = _case_mask_path(case_id, low=low)
        if not (os.path.exists(ct_path) and os.path.exists(mask_path)):
            return jsonify({"error": "CT or segmentation not found for this case."}), 404

        ct_obj = nib.load(ct_path)
        # float32 (see interactive_segment): halves the per-request RAM footprint.
        ct = ct_obj.get_fdata(dtype=np.float32)
        labels = nib.load(mask_path).get_fdata(dtype=np.float32)
        vessel_mask = (np.round(labels) == vessel_label).astype(np.uint8)
        if vessel_mask.sum() == 0:
            return jsonify({"error": "That vessel isn't segmented in this case."}), 422
        lesion_mask = (np.round(labels) == lesion_label).astype(np.uint8)
        has_lesion = lesion_mask.sum() > 0

        res = analyze_vessel(
            ct, ct_obj.affine, vessel_mask,
            lesion_mask if has_lesion else None,
            slab_radius_mm=float(body.get("slab_radius_mm", 20.0)),
        )

        # Window the reformat (default soft-tissue) and PNG-encode it.
        w = body.get("window") or {}
        width = float(w.get("width", 400)); center = float(w.get("center", 40))
        lo, hi = center - width / 2, center + width / 2
        img = np.clip((res["reformat"] - lo) / max(hi - lo, 1e-6), 0, 1)
        img8 = (img * 255).astype(np.uint8)

        from PIL import Image
        png = io.BytesIO()
        Image.fromarray(img8, mode="L").save(png, format="PNG")
        import base64
        data_url = "data:image/png;base64," + base64.b64encode(png.getvalue()).decode()

        return jsonify({
            "length_mm": round(res["length_mm"], 1),
            "max_contact_deg": round(res["max_contact_deg"], 1),
            "contact_length_mm": round(res["contact_length_mm"], 1),
            "has_lesion": bool(has_lesion),
            "num_points": res["num_points"],
            "contact_profile": [round(v, 1) for v in res["contact_profile"]],
            "reformat_png": data_url,
            "reformat_size": list(res["reformat"].shape),
        })
    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400
    except Exception as error:
        print("[vessel_cpr error]", type(error).__name__, error)
        return jsonify({"error": "Vessel analysis failed."}), 500
    finally:
        _ANALYSIS_SLOTS.release()
